"""Validate Excel file by loading with openpyxl"""
from openpyxl import load_workbook

print("Loading Excel file...")
try:
    wb = load_workbook('BSP/Q3-2025.xlsx', data_only=False)
    print(f"Successfully loaded workbook with {len(wb.sheetnames)} sheets")

    # Check first employee sheet (should be sheet 3 - after Budget and Cover)
    sheet_names = wb.sheetnames
    print(f"\nSheet names: {sheet_names[:5]}...")

    # Check a specific employee sheet
    if len(sheet_names) >= 3:
        ws = wb[sheet_names[2]]  # Third sheet (first employee)
        print(f"\nChecking sheet: {sheet_names[2]}")

        # Check row 7 formulas
        print(f"\nRow 7 formulas:")
        print(f"  L7 (Stundensatz): {ws['L7'].value[:100] if ws['L7'].value else 'None'}...")
        print(f"  N7 (Budget): {ws['N7'].value}")

        # Check row 62
        print(f"\nRow 62 formulas:")
        print(f"  L62 (Stundensatz): {ws['L62'].value[:100] if ws['L62'].value else 'None'}...")
        print(f"  N62 (Budget): {ws['N62'].value}")

        # Check max column
        max_col = ws.max_column
        print(f"\nMax column in sheet: {max_col}")

    print("\nExcel file is valid!")

except Exception as e:
    print(f"\nError loading Excel: {e}")
    import traceback
    traceback.print_exc()
