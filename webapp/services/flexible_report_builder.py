# -*- coding: utf-8 -*-
"""
Builder for flexible, non-quarterly reports.

This module is separated from the main report_generator.py to ensure that
the complex logic for standard quarterly reports is not affected by changes
for flexible reports.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from openpyxl.worksheet.datavalidation import DataValidation

from ..models import ReportConfig, ReportType, TimeBlock
from ..report_generator import (
    MONTHLY_BUDGETS,
    QUARTERLY_BUDGETS,
    _create_project_budget_sheet,
    _add_vba_macro,
    de_to_float,
    detect_billing_type,
    is_bonus_project,
    norm_ms,
    status_color_hex,
    ProgressCallback,
    _noop_progress,
)


def build_flexible_report(
    config: ReportConfig,
    df_csv: pd.DataFrame,
    df_budget: pd.DataFrame,
    milestone_parent_map: Dict,
    time_blocks: List[TimeBlock],
    out_path: Path,
    progress_cb: ProgressCallback = _noop_progress,
    add_vba: bool = True,
) -> Path:
    """
    Builds a flexible Excel report based on custom time blocks and filters.
    This function is designed to be safe and isolated from the main quarterly report logic.
    """
    if config.report_type in [ReportType.PROJECT_SUMMARY, ReportType.EMPLOYEE_SUMMARY]:
        return _build_summary_report(config, time_blocks, out_path)

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if config.include_budget_overview:
        progress_cb(18, "Erstelle Projekt-Budget-Übersicht")
        _create_project_budget_sheet(wb, df_budget, border)

    all_data = pd.concat([block.data for block in time_blocks])
    if config.exclude_special_projects:
        all_data = all_data[~all_data['proj_norm'].apply(is_bonus_project)]

    employees = sorted(all_data["staff_name"].unique())
    total_emps = max(len(employees), 1)

    # Precompute cumulative hours per (staff_name, proj_norm, ms_norm, quarter) for
    # quarterly-budget milestones (e.g. "Firmenveranstaltungen max. 4h/Quartal").
    # This lets us evaluate bonus eligibility per quarter even when time blocks are months.
    _all_wq = all_data.copy()
    _all_wq['_quarter_period'] = _all_wq['date_parsed'].dt.to_period('Q')
    quarterly_cum_map: dict = (
        _all_wq
        .groupby(['staff_name', 'proj_norm', 'ms_norm', '_quarter_period'])['hours']
        .sum()
        .to_dict()
    )

    employee_summary_data = {}

    for idx_emp, emp in enumerate(employees, start=1):
        sheet_name = emp[:31]
        ws = wb.create_sheet(title=sheet_name)
        employee_summary_data[emp] = {'sheet_name': sheet_name, 'blocks': {}}

        ws.append([f"{emp} - Report für {config.start_date.strftime('%d.%m.%Y')} - {config.end_date.strftime('%d.%m.%Y')}"])
        ws.append(["Position:", "SV"])
        ws.append([])
        current_row = 4

        for time_block in time_blocks:
            df_block_data = time_block.data[time_block.data["staff_name"] == emp].copy()
            if df_block_data.empty:
                continue

            block_hours = df_block_data.groupby(['proj_norm', 'ms_norm'], as_index=False).agg({'hours': 'sum'})
            block_data_merged = block_hours.merge(
                df_csv, how="left", on=["proj_norm", "ms_norm"]
            )
            block_data_merged["Projekte"] = block_data_merged["Projekte"].fillna(block_data_merged["proj_norm"])
            block_data_merged["Meilenstein"] = block_data_merged["Meilenstein"].fillna(block_data_merged["ms_norm"])

            block_quarter = pd.Period(time_block.start, freq='Q')

            for idx, row in block_data_merged.iterrows():
                ms_name = row["Meilenstein"]
                proj_name = row.get("Projekte", "")
                if is_bonus_project(proj_name) and ms_name in MONTHLY_BUDGETS:
                    full_month_budget = MONTHLY_BUDGETS[ms_name]
                    block_period = pd.Period(time_block.start, freq='M')
                    days_in_month = block_period.days_in_month
                    if days_in_month > 0:
                        prorated_soll = (time_block.duration_days / days_in_month) * full_month_budget
                        block_data_merged.loc[idx, "Soll"] = prorated_soll
                    else:
                        block_data_merged.loc[idx, "Soll"] = 0
                    block_data_merged.loc[idx, "Ist"] = row["hours"]
                elif is_bonus_project(proj_name) and ms_name in QUARTERLY_BUDGETS:
                    # Quarterly budget: Soll = fixed quarterly limit,
                    # Ist = cumulative hours within the same quarter (for % / bonus eligibility).
                    # "Stunden in Block" (hours_val) stays as the actual block hours.
                    quarterly_budget = QUARTERLY_BUDGETS[ms_name]
                    cum_q = quarterly_cum_map.get(
                        (emp, row["proj_norm"], row["ms_norm"], block_quarter), 0.0
                    )
                    block_data_merged.loc[idx, "Soll"] = quarterly_budget
                    block_data_merged.loc[idx, "Ist"]  = cum_q

            ws.append([f"--- {time_block.name} ---"])
            ws[f"A{current_row}"].font = Font(bold=True, size=12)
            current_row += 1

            # Header mit Bonus-Anpassung (intern) und Abrechnungsart
            header = [
                "Projekt",              # A (1)
                "Meilenstein",          # B (2)
                "Soll (h)",             # C (3)
                "Ist (h)",              # D (4)
                "Stunden in Block (h)", # E (5)
                "%",                    # F (6)
                "Bonus-Anpassung (h)",  # G (7) - nur intern, nicht für Export
                "Abrechnungsart",       # H (8) - neu, wird exportiert
                "Rechnung",             # I (9) - wird exportiert
                "Kommentar",            # J (10) - wird exportiert
            ]
            ws.append(header)

            for cell in ws[current_row]:
                cell.font = Font(bold=True); cell.border = border
            current_row += 1
            
            block_data_start_row = current_row
            bonus_hours_block = 0.0
            bonus_hours_special_block = 0.0
            adjustment_cells = []

            for _, row_data in block_data_merged.sort_values(["Projekte", "Meilenstein"]).iterrows():
                soll_val = row_data.get("Soll", 0.0) or 0.0
                ist_val = row_data.get("Ist", 0.0) or 0.0
                hours_val = row_data.get("hours", 0.0) or 0.0

                prozent = (ist_val / soll_val * 100.0) if soll_val > 0 else 0.0

                # Abrechnungsart ermitteln
                arbeitspaket = row_data.get("Arbeitspaket", "")
                honorarbereich = row_data.get("Honorarbereich", "")
                billing_type = detect_billing_type(arbeitspaket, honorarbereich)

                # Erstelle Zeile mit allen Spalten
                row_to_append = [
                    row_data["Projekte"],    # A (1)
                    row_data["Meilenstein"], # B (2)
                    round(soll_val, 2) if soll_val > 0 else "-",  # C (3)
                    round(ist_val, 2) if ist_val > 0 else "-",    # D (4)
                    round(hours_val, 2),     # E (5)
                    round(prozent, 2) if soll_val > 0 else "-",   # F (6)
                    "",                      # G (7) Bonus-Anpassung (editierbar, nur intern)
                    billing_type,            # H (8) Abrechnungsart (wird exportiert)
                    "",                      # I (9) Rechnung (Dropdown, wird exportiert)
                    "",                      # J (10) Kommentar (wird exportiert)
                ]
                ws.append(row_to_append)
                
                # Bonus-Berechnung
                is_special = is_bonus_project(row_data.get("Projekte", ""))
                bonus_candidate = prozent <= 100.0

                # Spalte G (7) = Bonus-Anpassung (editierbar, nur intern)
                adj_cell = ws.cell(row=current_row, column=7)
                adj_cell.number_format = "0.00"
                adjustment_cells.append(adj_cell.coordinate)

                if bonus_candidate:
                    if is_special:
                        bonus_hours_special_block += hours_val
                    else:
                        bonus_hours_block += hours_val

                # Spalte I (9) = Rechnung Dropdown
                rechnung_cell = ws.cell(row=current_row, column=9)
                rechnung_dv = DataValidation(type="list", formula1='"SR,AZ"', allow_blank=True)
                rechnung_dv.add(rechnung_cell)
                ws.add_data_validation(rechnung_dv)

                if soll_val > 0:
                    pct_cell = ws.cell(row=current_row, column=6)
                    pct_cell.fill = PatternFill(start_color=status_color_hex(prozent), end_color=status_color_hex(prozent), fill_type="solid")

                for cell in ws[current_row]:
                    cell.border = border
                current_row += 1
            
            block_data_end_row = current_row - 1

            sum_formula = f"=SUM(E{block_data_start_row}:E{block_data_end_row})"
            ws.append(["", "Summe", "", "", sum_formula])
            sum_total_cell = ws.cell(row=current_row, column=5)
            for cell in ws[current_row]: cell.font = Font(bold=True)
            sum_total_cell.number_format = "0.00"
            current_row += 1

            block_summary = {'total_hours_cell': sum_total_cell.coordinate}

            if config.include_bonus_calc:
                adj_sum_part = f",{','.join(adjustment_cells)}" if adjustment_cells else ""
                bonus_total_formula = f"=SUM({round(bonus_hours_block, 2)}{adj_sum_part})"
                ws.append(["", "Bonusberechtigte Stunden", "", "", bonus_total_formula])
                bonus_total_cell = ws.cell(row=current_row, column=5)
                for cell in ws[current_row]: cell.font = Font(bold=True)
                bonus_total_cell.number_format = "0.00"
                current_row += 1
                block_summary['bonus_hours_cell'] = bonus_total_cell.coordinate

                ws.append(["", "Bonusberechtigte Stunden Sonderprojekt", "", "", round(bonus_hours_special_block, 2)])
                special_bonus_cell = ws.cell(row=current_row, column=5)
                for cell in ws[current_row]: cell.font = Font(bold=True)
                special_bonus_cell.number_format = "0.00"
                current_row += 1
                block_summary['special_bonus_hours_cell'] = special_bonus_cell.coordinate
            
            employee_summary_data[emp]['blocks'][time_block.name] = block_summary

            ws.append([])
            current_row += 1

        # Spaltenbreiten optimiert (nicht zu breit)
        ws.column_dimensions['A'].width = 35  # Projekt
        ws.column_dimensions['B'].width = 45  # Meilenstein
        ws.column_dimensions['C'].width = 10  # Soll
        ws.column_dimensions['D'].width = 10  # Ist
        ws.column_dimensions['E'].width = 15  # Stunden in Block
        ws.column_dimensions['F'].width = 8   # %
        ws.column_dimensions['G'].width = 15  # Bonus-Anpassung (nur intern)
        ws.column_dimensions['H'].width = 15  # Abrechnungsart
        ws.column_dimensions['I'].width = 12  # Rechnung
        ws.column_dimensions['J'].width = 25  # Kommentar

        progress = int((idx_emp / total_emps) * 80) + 20
        progress_cb(min(progress, 95), f"Verarbeite Mitarbeiter {emp}")

    if config.include_summary_sheet:
        _create_flexible_summary_sheet(wb, employee_summary_data, time_blocks, config)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    # Add VBA macro if requested and convert to .xlsm
    if add_vba:
        progress_cb(98, "Füge VBA-Makro hinzu")
        final_path = _add_vba_macro(out_path, progress_cb)
    else:
        final_path = out_path

    progress_cb(100, "Flexibler Report fertiggestellt")
    return final_path


def _create_flexible_summary_sheet(
    wb: Workbook,
    employee_summary_data: Dict,
    time_blocks: List[TimeBlock],
    config: ReportConfig,
) -> None:
    """Creates a summary sheet for a flexible report."""
    ws = wb.create_sheet(title="Übersicht", index=0)
    border = Border(left=Side(style="thin", color="DDDDDD"),
                    right=Side(style="thin", color="DDDDDD"),
                    top=Side(style="thin", color="DDDDDD"),
                    bottom=Side(style="thin", color="DDDDDD"))

    title = f"Zusammenfassung für {config.start_date.strftime('%d.%m.%Y')} - {config.end_date.strftime('%d.%m.%Y')}"
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    current_row = 3

    ws.append(["--- Summen pro Zeit-Block ---"])
    ws[f"A{current_row}"].font = Font(bold=True, size=12)
    current_row += 1

    header = ["Zeit-Block", "Gesamtstunden"]
    if config.include_bonus_calc:
        header.extend(["Bonusberechtigte Stunden", "Bonusberechtigte Stunden Sonderprojekt"])
    ws.append(header)
    for cell in ws[current_row]:
        cell.font = Font(bold=True)
        cell.border = border
    current_row += 1

    summary_start_row = current_row
    for block in time_blocks:
        total_hours_refs = []
        bonus_hours_refs = []
        special_bonus_refs = []

        for emp, emp_data in employee_summary_data.items():
            if block.name in emp_data['blocks']:
                block_data = emp_data['blocks'][block.name]
                sheet_name = emp_data['sheet_name']
                total_hours_refs.append(f"'{sheet_name}'!{block_data['total_hours_cell']}")
                if config.include_bonus_calc and 'bonus_hours_cell' in block_data:
                    bonus_hours_refs.append(f"'{sheet_name}'!{block_data['bonus_hours_cell']}")
                if config.include_bonus_calc and 'special_bonus_hours_cell' in block_data:
                    special_bonus_refs.append(f"'{sheet_name}'!{block_data['special_bonus_hours_cell']}")
        
        row_to_append = [block.name]
        row_to_append.append(f"=SUM({','.join(total_hours_refs)})" if total_hours_refs else 0)
        if config.include_bonus_calc:
            row_to_append.append(f"=SUM({','.join(bonus_hours_refs)})" if bonus_hours_refs else 0)
            row_to_append.append(f"=SUM({','.join(special_bonus_refs)})" if special_bonus_refs else 0)
        
        ws.append(row_to_append)
        for cell in ws[current_row]:
            cell.border = border
            if cell.column > 1:
                cell.number_format = "0.00"
        current_row += 1
    summary_end_row = current_row - 1

    ws.append([])
    current_row += 1

    # --- Grand Totals ---
    ws.append(["--- Gesamtsumme ---"])
    ws[f"A{current_row}"].font = Font(bold=True, size=12)
    current_row += 1

    total_hours_formula = f"=SUM(B{summary_start_row}:B{summary_end_row})"
    ws.append(["Gesamt eingetragene Stunden:", total_hours_formula])
    ws[f"B{current_row}"].number_format = "0.00"
    for cell in ws[current_row]: cell.font = Font(bold=True); cell.border = border
    current_row += 1

    if config.include_bonus_calc:
        total_bonus_formula = f"=SUM(C{summary_start_row}:C{summary_end_row})"
        ws.append(["Bonusberechtigte Stunden (Gesamt):", total_bonus_formula])
        ws[f"B{current_row}"].number_format = "0.00"
        for cell in ws[current_row]: cell.font = Font(bold=True); cell.border = border
        current_row += 1

        total_special_bonus_formula = f"=SUM(D{summary_start_row}:D{summary_end_row})"
        ws.append(["Bonusberechtigte Stunden Sonderprojekt (Gesamt):", total_special_bonus_formula])
        ws[f"B{current_row}"].number_format = "0.00"
        for cell in ws[current_row]: cell.font = Font(bold=True); cell.border = border
        current_row += 1

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 25
    if config.include_bonus_calc:
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 40



def _build_summary_report(
    config: ReportConfig,
    time_blocks: List[TimeBlock],
    out_path: Path,
) -> Path:
    """
    Builds a simplified summary report for projects or employees.
    This is a placeholder implementation.
    """
    wb = Workbook()
    ws = wb.active

    if config.report_type == ReportType.PROJECT_SUMMARY:
        ws.title = "Projekt-Zusammenfassung"
        ws.append(["Projekt-Zusammenfassung"])
        ws.append([f"Zeitraum: {config.start_date.strftime('%d.%m.%Y')} - {config.end_date.strftime('%d.%m.%Y')}"])
        ws.append([])
        ws.append(["Projekt", "Mitarbeiter", "Stunden"])
        
        all_data = pd.concat([block.data for block in time_blocks])
        summary = all_data.groupby(['proj_norm', 'staff_name'])['hours'].sum().reset_index()

        for _, row in summary.iterrows():
            ws.append([row['proj_norm'], row['staff_name'], row['hours']])

    elif config.report_type == ReportType.EMPLOYEE_SUMMARY:
        ws.title = "Mitarbeiter-Zusammenfassung"
        ws.append(["Mitarbeiter-Zusammenfassung"])
        ws.append([f"Zeitraum: {config.start_date.strftime('%d.%m.%Y')} - {config.end_date.strftime('%d.%m.%Y')}"])
        ws.append([])
        ws.append(["Mitarbeiter", "Projekt", "Stunden"])

        all_data = pd.concat([block.data for block in time_blocks])
        summary = all_data.groupby(['staff_name', 'proj_norm'])['hours'].sum().reset_index()

        for _, row in summary.iterrows():
            ws.append([row['staff_name'], row['proj_norm'], row['hours']])

    wb.save(out_path)
    return out_path
