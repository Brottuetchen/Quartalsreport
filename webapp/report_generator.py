# -*- coding: utf-8 -*-
"""Core logic for generating quarterly bonus reports.

This module refactors the original CLI script into reusable functions that can
be imported by a web server or other automation. The main entry point is
`generate_quarterly_report`, which expects paths to the CSV and XML sources and
returns the path to the generated Excel workbook.
"""

from __future__ import annotations

import math
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Set, Tuple

import numpy as np
import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

# ===================== BUDGETS FÜR 0000-PROJEKT =====================
# Diese Budgets gelten für ALLE Mitarbeiter, die diese Meilensteine bearbeiten
# Format: {"Meilensteinname": Stunden_pro_Monat/Quartal}

# Monatliche Budgets (0000-Meilensteine mit "Monat" im Namen)
MONTHLY_BUDGETS: Dict[str, float] = {
    "Einarbeitung neuer Mitarbeiter (max. 8h/Monat pro MA)": 8.0,
    "Angebote-Ausschreibungen-Kalkulationen (max. 8h/Monat pro MA)": 8.0,
    "Erstellung Vorlagen (übergreifend) (max. 8h/Monat pro MA)": 8.0,
}

# Quartalsbudgets (0000-Meilensteine mit "Quartal" im Namen)
QUARTERLY_BUDGETS: Dict[str, float] = {
    "Firmenveranstaltungen (max. 4h/Quartal pro MA)": 4.0,
    "Vorträge, Repräsentation (übergreifend) (max. 4h/Quartal pro MA)": 4.0,
    "Messeauftritt (max. 4h/Quartal pro MA)": 4.0,
}


MONTH_NAMES = {
    1: "Januar",
    2: "Februar",
    3: "März",
    4: "April",
    5: "Mai",
    6: "Juni",
    7: "Juli",
    8: "August",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Dezember",
}

ProgressCallback = Callable[[int, str], None]


def _noop_progress(_: int, __: str) -> None:
    """Default progress callback that swallows updates."""


# ===================== Hilfsfunktionen =====================
def de_to_float(x):
    if pd.isna(x):
        return np.nan
    s = str(x).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return np.nan


def norm_ms(text: str) -> str:
    if text is None or (isinstance(text, float) and math.isnan(text)):
        return ""
    s = str(text).replace("\u2022", "").replace("•", "").replace("●", "")
    s = re.sub(r"^[\-\s]+", "", s)
    return s.strip()


def get_milestone_type(milestone_name: str) -> str:
    if milestone_name is None or (isinstance(milestone_name, float) and math.isnan(milestone_name)):
        return "monthly"
    name_lower = str(milestone_name).lower()
    return "quarterly" if "quartal" in name_lower else "monthly"


def extract_budget_from_name(ms_name):
    """Extrahiert Budgetstunden und Einheit (Monat/Quartal) aus dem Meilenstein-Namen."""

    if ms_name is None or (isinstance(ms_name, float) and math.isnan(ms_name)):
        return None, None
    text = str(ms_name)
    m = re.search(r"(?i)(\d+[\.,]?\d*)\s*h\s*(?:/|pro\s+)(monat|quartal)", text)
    if not m:
        return None, None
    try:
        hours = float(m.group(1).replace(',', '.'))
    except Exception:
        return None, None
    unit = m.group(2).lower()
    return hours, unit


def is_bonus_project(name: str) -> bool:
    if name is None or (isinstance(name, float) and math.isnan(name)):
        return False
    s = str(name).strip().lower()
    return s.startswith('0000')


def is_nachtrag_package(name: str) -> bool:
    if name is None or (isinstance(name, float) and math.isnan(name)):
        return False
    s = str(name).lower()
    return "nat" in s or "nachtrag" in s


def status_color_hex(p: float) -> str:
    if p < 90:
        return "C6EFCE"  # grün
    elif p <= 100:
        return "FFF2CC"  # gelb
    else:
        return "F8CBAD"  # rot


def detect_billing_type(arbeitspaket: str, honorarbereich: str, force: bool = False) -> str:
    """
    Erkennt die Abrechnungsart eines Projekts/Meilensteins.
    Returns: "Pauschale" | "Nachweis" | "Unbekannt"
    """
    if not force and (pd.isna(honorarbereich) or str(honorarbereich).strip().upper() != "X"):
        return "Unbekannt"  # Keine Obermeilenstein-Markierung

    if pd.isna(arbeitspaket):
        return "Unbekannt"

    arbeitspaket_lower = str(arbeitspaket).lower()

    # Pauschale erkennen: (p)
    if "(p)" in arbeitspaket_lower:
        return "Pauschale"

    # Nachweis erkennen: (aN), (a.N.), (a N)
    if any(marker in arbeitspaket_lower for marker in ["(an)", "(a.n.)", "(a n)"]):
        return "Nachweis"

    # Nicht eindeutig erkennbar
    return "Unbekannt"


def load_csv_budget_data(csv_path: Path) -> Tuple[pd.DataFrame, Dict[Tuple[str, str], Set[str]]]:
    """
    Lädt Budget-Informationen aus CSV für Projekt-Budget-Übersicht.
    Returns:
        Tuple[pd.DataFrame, Dict]: DataFrame mit Budgetdaten je Obermeilenstein sowie
        ein Mapping {(Projekt|Projektcode, Meilenstein): {Obermeilensteine}} für Zuordnungen.
    """
    try_encodings = [("utf-16", "\t"), ("utf-8-sig", "\t"), ("cp1252", "\t")]
    df = None
    for enc, delim in try_encodings:
        try:
            df = pd.read_csv(csv_path, delimiter=delim, encoding=enc)
            break
        except Exception:
            continue
    if df is None:
        raise RuntimeError("CSV konnte nicht gelesen werden.")

    df.columns = [c.strip().replace("\u200b", "").replace("\ufeff", "") for c in df.columns]

    # Check for 'Projekte' column or alternatives
    if "Projekte" not in df.columns:
        # Try to find alternative names
        found = False
        for alt in ["Projekt", "Project", "Projects", "Projektname"]:
            if alt in df.columns:
                df.rename(columns={alt: "Projekte"}, inplace=True)
                found = True
                break
        if not found:
            raise ValueError(f"Spalte 'Projekte' nicht gefunden. Verfügbare Spalten: {list(df.columns)}")

    df["Projekte"] = df["Projekte"].ffill()

    def _project_keys(name: str) -> List[str]:
        """Returns possible lookup keys for a project (full string + first token/code)."""
        if name is None:
            return []
        base = str(name).strip()
        if not base:
            return []
        keys = [base]
        first = base.split(maxsplit=1)[0].strip() if base.split() else base
        if first and first not in keys:
            keys.append(first)
        return keys

    milestone_parent_map: Dict[Tuple[str, str], Set[str]] = {}
    current_project = None
    current_parent_norm = None

    for _, raw_row in df.iterrows():
        projekt = str(raw_row.get("Projekte", "")).strip()
        arbeitspaket_raw = str(raw_row.get("Arbeitspaket", "")).strip()
        honorarbereich = str(raw_row.get("Honorarbereich", "")).strip().upper()

        if projekt != current_project:
            current_project = projekt
            current_parent_norm = None

        if not arbeitspaket_raw or arbeitspaket_raw == "-":
            current_parent_norm = None
            continue

        ms_norm_value = norm_ms(arbeitspaket_raw)
        if not ms_norm_value:
            continue

        project_keys = _project_keys(projekt)
        if honorarbereich == "X":
            current_parent_norm = ms_norm_value
            for key in project_keys:
                milestone_parent_map.setdefault((key, ms_norm_value), set()).add(ms_norm_value)
            continue

        if current_parent_norm:
            for key in project_keys:
                milestone_parent_map.setdefault((key, ms_norm_value), set()).add(current_parent_norm)

    # Nur Obermeilensteine (X-Markierung) interessieren uns für Budget-Übersicht
    mask_obermeilenstein = (
        df["Honorarbereich"].notna() &
        (df["Honorarbereich"].astype(str).str.strip().str.upper() == "X") &
        df["Arbeitspaket"].notna() &
        (df["Arbeitspaket"].astype(str).str.strip() != "-")
    )

    budget_rows = []
    added_budget_keys: Set[Tuple[str, str]] = set()

    for idx, row in df[mask_obermeilenstein].iterrows():
        projekt = str(row["Projekte"]).strip()
        arbeitspaket = str(row["Arbeitspaket"]).strip()
        projekt_code = projekt.split(maxsplit=1)[0].strip() if projekt.split() else projekt
        ober_norm = norm_ms(arbeitspaket)

        # Abrechnungsart erkennen
        billing_type = detect_billing_type(arbeitspaket, row["Honorarbereich"])

        # Budget-Daten extrahieren
        sollhonor = de_to_float(row.get("Sollhonorar", 0))
        verrechnete_honorare = de_to_float(row.get("Verrechnete Honorare", 0))
        istkosten = de_to_float(row.get("Istkosten", 0))
        sollstunden = de_to_float(row.get("Sollstunden Budget", 0))
        iststunden = de_to_float(row.get("Iststunden", 0))
        budget = de_to_float(row.get("Budget", 0))

        # Stundensätze für Positionen sammeln (aus Unterpositionen)
        rate_sv = None
        rate_cad = None
        rate_adm = None

        # Sollstunden von Untermeilensteinen aufsummieren
        total_sub_sollstunden = 0.0

        # Suche nach Unterpositionen mit SV/CAD/ADM
        # Nächste Zeilen nach dem Obermeilenstein durchsuchen
        for sub_idx in range(idx + 1, min(idx + 80, len(df))):
            sub_row = df.iloc[sub_idx]
            sub_arbeitspaket = str(sub_row.get("Arbeitspaket", "")).strip()
            if not sub_arbeitspaket or sub_arbeitspaket == "-":
                continue

            sub_honorar = str(sub_row.get("Honorarbereich", "")).strip().upper()
            if sub_honorar == "X":
                break  # nächster Obermeilenstein erreicht

            # Sollstunden aufsummieren (für alle Untermeilensteine)
            sub_sollstunden_val = de_to_float(sub_row.get("Sollstunden Budget", 0))
            if sub_sollstunden_val > 0:
                total_sub_sollstunden += sub_sollstunden_val

            # Extrahiere Position
            if "'   SV" in sub_arbeitspaket or "'   S V" in sub_arbeitspaket:
                sub_budget = de_to_float(sub_row.get("Budget", 0))
                sub_sollstunden = de_to_float(sub_row.get("Sollstunden Budget", 0))
                if sub_sollstunden > 0:
                    rate_sv = sub_budget / sub_sollstunden
            elif "'   CAD" in sub_arbeitspaket or "'   C A D" in sub_arbeitspaket:
                sub_budget = de_to_float(sub_row.get("Budget", 0))
                sub_sollstunden = de_to_float(sub_row.get("Sollstunden Budget", 0))
                if sub_sollstunden > 0:
                    rate_cad = sub_budget / sub_sollstunden
            elif "'   ADM" in sub_arbeitspaket or "'   A D M" in sub_arbeitspaket:
                sub_budget = de_to_float(sub_row.get("Budget", 0))
                sub_sollstunden = de_to_float(sub_row.get("Sollstunden Budget", 0))
                if sub_sollstunden > 0:
                    rate_adm = sub_budget / sub_sollstunden

            # Untermeilensteine mit eigenem Budget (z. B. NAT) separat aufnehmen
            if is_nachtrag_package(sub_arbeitspaket):
                sub_sollhonor = de_to_float(sub_row.get("Budget", 0))
                if pd.isna(sub_sollhonor) or sub_sollhonor == 0:
                    sub_sollhonor = de_to_float(sub_row.get("Sollhonorar", 0))
                if pd.isna(sub_sollhonor) or sub_sollhonor == 0:
                    continue
                sub_sollstunden = de_to_float(sub_row.get("Sollstunden Budget", 0))
                sub_iststunden = de_to_float(sub_row.get("Iststunden", 0))
                sub_verrechnete = de_to_float(sub_row.get("Verrechnete Honorare", 0))
                sub_istkosten = de_to_float(sub_row.get("Istkosten", 0))
                sub_norm = norm_ms(sub_arbeitspaket)
                sub_key = (projekt, sub_norm)
                if sub_key in added_budget_keys:
                    continue

                sub_billing = detect_billing_type(sub_arbeitspaket, "X", force=True)
                if sub_billing == "Unbekannt":
                    sub_billing = billing_type

                default_rate = sub_sollhonor / sub_sollstunden if sub_sollstunden and sub_sollstunden > 0 else None
                budget_rows.append({
                    "Projekt": projekt,
                    "ProjektCode": projekt_code,
                    "Obermeilenstein": sub_arbeitspaket,
                    "Obermeilenstein_norm": sub_norm,
                    "Abrechnungsart": sub_billing,
                    "Gesamtbudget": sub_sollhonor,
                    "Abgerechnet": sub_verrechnete,
                    "Istkosten": sub_istkosten,
                    "Sollstunden": sub_sollstunden,
                    "Iststunden": sub_iststunden,
                    "Stundensatz_SV": default_rate,
                    "Stundensatz_CAD": default_rate,
                    "Stundensatz_ADM": default_rate,
                    "LookupKey": f"{projekt}||{sub_norm}",
                })
                added_budget_keys.add(sub_key)
                project_keys = _project_keys(projekt)
                for key in project_keys:
                    milestone_parent_map.setdefault((key, sub_norm), set()).add(sub_norm)

        # Bei Pauschale: Berechne Stundensatz aus Sollhonor / Sollstunden
        if billing_type == "Pauschale" and sollstunden > 0:
            default_rate = sollhonor / sollstunden
            if rate_sv is None:
                rate_sv = default_rate
            if rate_cad is None:
                rate_cad = default_rate
            if rate_adm is None:
                rate_adm = default_rate

        effective_sollstunden = total_sub_sollstunden if total_sub_sollstunden > 0 else sollstunden

        main_key = (projekt, ober_norm)
        if main_key not in added_budget_keys:
            budget_rows.append({
                "Projekt": projekt,
                "ProjektCode": projekt_code,
                "Obermeilenstein": arbeitspaket,
                "Obermeilenstein_norm": ober_norm,
                "Abrechnungsart": billing_type,
                "Gesamtbudget": sollhonor,
                "Abgerechnet": verrechnete_honorare,
                "Istkosten": istkosten,
                "Sollstunden": effective_sollstunden,
                "Iststunden": iststunden,
                "Stundensatz_SV": rate_sv,
                "Stundensatz_CAD": rate_cad,
                "Stundensatz_ADM": rate_adm,
                "LookupKey": f"{projekt}||{ober_norm}",
            })
            added_budget_keys.add(main_key)

    df_budget = pd.DataFrame(budget_rows)
    df_budget["_LookupId"] = range(1, len(df_budget) + 1)
    return df_budget, milestone_parent_map


def load_csv_projects(csv_path: Path) -> pd.DataFrame:
    """CSV laden (Soll/Ist-Basis)."""

    try_encodings = [("utf-16", "\t"), ("utf-8-sig", "\t"), ("cp1252", "\t")]
    df = None
    for enc, delim in try_encodings:
        try:
            df = pd.read_csv(csv_path, delimiter=delim, encoding=enc)
            break
        except Exception:
            continue
    if df is None:
        raise RuntimeError("CSV konnte nicht gelesen werden.")

    df.columns = [c.strip().replace("\u200b", "").replace("\ufeff", "") for c in df.columns]

    # Check for 'Projekte' column or alternatives
    if "Projekte" not in df.columns:
        # Try to find alternative names
        found = False
        for alt in ["Projekt", "Project", "Projects", "Projektname"]:
            if alt in df.columns:
                df.rename(columns={alt: "Projekte"}, inplace=True)
                found = True
                break
        if not found:
            raise ValueError(f"Spalte 'Projekte' nicht gefunden. Verfügbare Spalten: {list(df.columns)}")

    df["Projekte"] = df["Projekte"].ffill()

    mask_ms = df["Arbeitspaket"].notna() & (df["Arbeitspaket"].astype(str).str.strip() != "-")
    cols_need = ["Projekte", "Arbeitspaket", "Iststunden", "Sollstunden Budget"]
    ms = df.loc[mask_ms, cols_need].copy()
    ms["Ist"] = ms["Iststunden"].map(de_to_float)
    ms["Soll"] = ms["Sollstunden Budget"].map(de_to_float)
    ms["Meilenstein"] = ms["Arbeitspaket"].map(norm_ms)
    ms = ms[["Projekte", "Meilenstein", "Ist", "Soll"]]

    g = ms.groupby(["Projekte", "Meilenstein"], as_index=False).agg({"Soll": "sum", "Ist": "sum"})
    g["Prozent"] = np.where(
        g["Soll"] > 0,
        (g["Ist"] / g["Soll"]) * 100.0,
        np.where(g["Ist"] > 0, 999.0, 0.0),
    )
    g["proj_norm"] = g["Projekte"].astype(str).str.strip()
    g["ms_norm"] = g["Meilenstein"].map(norm_ms)
    return g


def load_xml_times(xml_path: Path) -> pd.DataFrame:
    """XML laden (Zeiteinträge)."""

    tree = ET.parse(xml_path)
    root = tree.getroot()
    rows = []

    for element in root.iter():
        if element.tag != "row":
            continue
        cells = element.findall("./cell")
        if not cells:
            continue
        entry = {cell.attrib.get("name"): (cell.text or "").strip() for cell in cells}
        if entry and "staff_name" in entry and entry["staff_name"] and "work_package_name" in entry and "date" in entry:
            rows.append(entry)

    if not rows:
        raise ValueError("XML enthält keine Daten.")
    df = pd.DataFrame(rows)

    def extract_date(date_str):
        match = re.search(r'(\d{1,2})\s+(\w{3})\s+(\d{4})', date_str)
        if match:
            day, month_str, year = match.groups()
            months = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06',
                      'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
            month = months.get(month_str, '01')
            month = months.get(month_str, '01')
            return f"{year}-{month}-{day.zfill(2)}"
        return None

    df["date_parsed"] = df["date"].apply(extract_date)
    df["date_parsed"] = pd.to_datetime(df["date_parsed"], errors='coerce')
    df["period"] = df["date_parsed"].dt.to_period("M")
    df["quarter"] = df["date_parsed"].dt.to_period("Q")
    df["proj_norm"] = df["project"].astype(str).str.strip()
    df["ms_norm"] = df["work_package_name"].map(norm_ms)

    def parse_hours(x):
        if pd.isna(x):
            return 0.0
        try:
            return float(str(x).strip())
        except Exception:
            return 0.0

    df['hours'] = df.get('number', 0).apply(parse_hours)
    return df


def list_available_quarters(df_xml: pd.DataFrame) -> Dict[pd.Period, List[pd.Period]]:
    """Gibt verfügbare Quartale und zugehörige Monate zurück."""

    quarters: Dict[pd.Period, List[pd.Period]] = {}
    for period in df_xml["period"].dropna().unique():
        quarter = period.to_timestamp().to_period("Q")
        quarters.setdefault(quarter, [])
        quarters[quarter].append(period)
    for quarter, months in quarters.items():
        months.sort()
    return quarters


def parse_quarter(quarter_str: str) -> pd.Period:
    """Parst Eingaben wie "2025Q3" oder "Q3-2025"."""

    quarter_str = quarter_str.strip().upper()
    match = re.match(r"^Q(\d)[-/]?\s*(\d{4})$", quarter_str)
    if match:
        q, year = match.groups()
        return pd.Period(year=int(year), quarter=int(q), freq="Q")
    match = re.match(r"^(\d{4})[-/\s]*Q(\d)$", quarter_str)
    if match:
        year, q = match.groups()
        return pd.Period(year=int(year), quarter=int(q), freq="Q")
    raise ValueError(f"Ungültiges Quartal: {quarter_str}")


@dataclass
class QuarterSelection:
    period: pd.Period
    months: List[pd.Period]


def determine_quarter(df_xml: pd.DataFrame, requested: Optional[str] = None) -> QuarterSelection:
    """Wählt das Zielquartal basierend auf den XML-Daten."""

    available = list_available_quarters(df_xml)
    if not available:
        raise ValueError("Keine Quartale in den XML-Daten gefunden.")

    if requested:
        target = parse_quarter(requested)
        if target not in available:
            raise ValueError(f"Angefordertes Quartal {requested} nicht in den XML-Daten enthalten.")
        months = sorted(available[target])
        return QuarterSelection(period=target, months=months)

    # Standard: jüngstes Quartal wählen
    target = sorted(available.keys())[-1]
    months = sorted(available[target])
    return QuarterSelection(period=target, months=months)


def _create_project_budget_sheet(
    wb: Workbook,
    df_budget: pd.DataFrame,
    border: Border,
) -> None:
    """Creates a project budget overview sheet with billing type and hourly rates."""

    if "Projekt-Budget-Übersicht" in wb.sheetnames:
        ws = wb["Projekt-Budget-Übersicht"]
        # Clear/Reset content
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(title="Projekt-Budget-Übersicht", index=0)

    # Title
    ws.append(["Projekt-Budget-Übersicht"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Hinweis: Rote Zellen = Manuelle Eingabe erforderlich | Gelbe Zellen = Optional manuell anpassen"])
    ws["A3"].font = Font(italic=True, size=10)
    ws.append([])

    current_row = 5
    header_row = current_row

    # Header row
    headers = [
        "Projekt",
        "Obermeilenstein",
        "Abrechnungsart",
        "Sollstunden Budget (h)",
        "Status",
        "Gesamtbudget (€)",
        "Abgerechnet (€)",
        "Verfügbar (€)",
        "Stundensatz SV (€/h)",
        "Stundensatz CAD (€/h)",
        "Stundensatz ADM (€/h)",
        "Bemerkung",
        "_LookupId",
    ]
    ws.append(headers)
    for cell in ws[current_row]:
        cell.font = Font(bold=True)
        cell.border = border
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
    current_row += 1

    data_start_row = current_row + 1

    # Data rows
    for _, row_data in df_budget.iterrows():
        projekt = row_data["Projekt"]
        obermeilenstein = row_data["Obermeilenstein"]
        billing_type = row_data["Abrechnungsart"]
        sollstunden = row_data.get("Sollstunden", 0)
        gesamtbudget = row_data["Gesamtbudget"]
        abgerechnet = row_data["Abgerechnet"]
        rate_sv = row_data["Stundensatz_SV"]
        rate_cad = row_data["Stundensatz_CAD"]
        rate_adm = row_data["Stundensatz_ADM"]

        # Status berechnen
        has_billing_type = billing_type != "Unbekannt"
        has_rates = (
            (rate_sv is not None and rate_sv > 0) or
            (rate_cad is not None and rate_cad > 0) or
            (rate_adm is not None and rate_adm > 0)
        )

        if has_billing_type and has_rates:
            status = "✓ Vollständig"
        else:
            status = "⚠ Manual erforderlich"

        # Verfügbar als Formel
        verfuegbar_formula = f"=F{current_row}-G{current_row}"

        # Bemerkung generieren
        bemerkung = []
        if not has_billing_type:
            bemerkung.append("Abrechnungsart wählen")
        if not has_rates:
            bemerkung.append("Stundensätze eingeben")
        bemerkung_text = "; ".join(bemerkung)

        ws.append([
            projekt,
            obermeilenstein,
            billing_type,
            sollstunden if sollstunden > 0 else "",
            status,
            gesamtbudget if gesamtbudget > 0 else "",
            abgerechnet if abgerechnet > 0 else 0,
            verfuegbar_formula,
            rate_sv if rate_sv is not None else "",
            rate_cad if rate_cad is not None else "",
            rate_adm if rate_adm is not None else "",
            bemerkung_text,
            row_data.get("_LookupId", ""),
        ])

        # Styling für die Zeile
        for col_idx, cell in enumerate(ws[current_row], start=1):
            cell.border = border

            # Abrechnungsart Dropdown (Spalte C)
            if col_idx == 3:
                dv = DataValidation(type="list", formula1='"Pauschale,Nachweis,Unbekannt"', allow_blank=False)
                dv.add(cell)
                ws.add_data_validation(dv)

                # Rot markieren wenn "Unbekannt"
                if billing_type == "Unbekannt":
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    cell.font = Font(color='9C0006', bold=True)

            # Sollstunden Spalte (Spalte D)
            if col_idx == 4:
                cell.number_format = '#,##0.00'

            # Status Spalte (Spalte E)
            if col_idx == 5:
                if status.startswith("⚠"):
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    cell.font = Font(color='9C5700', bold=True)
                else:
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    cell.font = Font(color='006100', bold=True)

            # Budget-Spalten (F, G, H)
            if col_idx in [6, 7, 8]:
                cell.number_format = '#,##0.00'

            # Stundensatz-Spalten (I, J, K) - Gelb markieren wenn leer
            if col_idx in [9, 10, 11]:
                cell.number_format = '#,##0.00'
                if cell.value == "":
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

            # Stundensatz-Spalten (I, J, K) - Gelb markieren wenn leer
            if col_idx in [8, 9, 10]:
                cell.number_format = '#,##0.00'
                if cell.value == "":
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

        current_row += 1

    # Enable filter row and freeze panes for easier navigation
    if current_row > data_start_row:
        ws.auto_filter.ref = f"A{header_row}:M{current_row - 1}"
    ws.freeze_panes = f"A{data_start_row}"

    # Column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 22
    ws.column_dimensions['J'].width = 22
    ws.column_dimensions['K'].width = 22
    ws.column_dimensions['L'].width = 40
    ws.column_dimensions['M'].width = 3
    ws.column_dimensions['M'].hidden = True


def _create_cover_sheet(
    wb: Workbook,
    target_quarter: pd.Period,
    months: Iterable[pd.Period],
    employee_summary_data: Dict,
    border: Border,
    report_title: Optional[str] = None,
) -> None:
    """Creates a cover sheet with summary totals across all employees."""

    # Create summary sheet
    # Create summary sheet
    if "Übersicht" in wb.sheetnames:
        ws = wb["Übersicht"]
        # Clear existing content if any (start from scratch but keep button)
        # Note: clearing cells might be safer than delete_rows if we want to preserve objects
        # But openpyxl delete_rows is usually fine. Let's start writing at row 1.
        # Clearing content manually:
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
                cell.font = Font()
    else:
        ws = wb.create_sheet(title="Übersicht", index=0)

    # Title
    title = report_title if report_title else f"Quartalsübersicht {target_quarter}"
    ws.append([f"{title} - Zusammenfassung aller Mitarbeiter"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    current_row = 3

    # Monthly summary table
    ws.append(["--- Monatliche Summen ---"])
    ws[f"A{current_row}"].font = Font(bold=True, size=12)
    current_row += 1

    # Header row
    ws.append(["Monat", "Gesamtstunden", "Bonusberechtigte Stunden", "Bonusberechtigte Stunden Sonderprojekt"])
    for cell in ws[current_row]:
        cell.font = Font(bold=True)
        cell.border = border
    current_row += 1

    # Build month labels from the authoritative `months` parameter so that every month
    # in the quarter appears in the cover sheet, even if the first employee has no
    # entries for that month.
    if employee_summary_data:
        month_labels = [
            f"{MONTH_NAMES.get(int(m.month), m.strftime('%B'))} {m.year}"
            for m in months
        ]

        # For each month, create a summary row
        for month_label in month_labels:
            # Collect cell references for all employees for this month
            total_hours_refs = []
            bonus_hours_refs = []
            special_bonus_refs = []

            for emp, emp_data in employee_summary_data.items():
                if month_label in emp_data['months']:
                    month_data = emp_data['months'][month_label]
                    total_hours_refs.append(month_data['total_hours_cell'])
                    bonus_hours_refs.append(month_data['bonus_hours_cell'])
                    special_bonus_refs.append(month_data['special_bonus_hours_cell'])

            # Create formulas summing across all employees
            total_hours_formula = f"=SUM({','.join(total_hours_refs)})" if total_hours_refs else "0"
            bonus_hours_formula = f"=SUM({','.join(bonus_hours_refs)})" if bonus_hours_refs else "0"
            special_bonus_formula = f"=SUM({','.join(special_bonus_refs)})" if special_bonus_refs else "0"

            ws.append([month_label, total_hours_formula, bonus_hours_formula, special_bonus_formula])
            for cell in ws[current_row]:
                cell.border = border
                if cell.column > 1:  # Format numeric columns
                    cell.number_format = "0.00"
            current_row += 1

    ws.append([])
    current_row += 1

    # Quarterly summary
    ws.append(["--- Quartalssummen ---"])
    ws[f"A{current_row}"].font = Font(bold=True, size=12)
    current_row += 1

    # Collect quarterly total cell references from all employees
    quarter_total_refs = []
    quarter_bonus_refs = []
    quarter_special_refs = []

    for emp, emp_data in employee_summary_data.items():
        if 'quarter_total_hours_cell' in emp_data:
            quarter_total_refs.append(emp_data['quarter_total_hours_cell'])
        if 'quarter_bonus_hours_cell' in emp_data:
            quarter_bonus_refs.append(emp_data['quarter_bonus_hours_cell'])
        if 'quarter_special_bonus_hours_cell' in emp_data:
            quarter_special_refs.append(emp_data['quarter_special_bonus_hours_cell'])

    # Total hours
    ws.append(["Gesamt eingetragene Stunden:", f"=SUM({','.join(quarter_total_refs)})" if quarter_total_refs else "0"])
    ws[f"B{current_row}"].number_format = "0.00"
    for cell in ws[current_row]:
        cell.font = Font(bold=True)
        cell.border = border
    current_row += 1

    # Bonus hours
    ws.append(["Bonusberechtigte Stunden (Quartal):", f"=SUM({','.join(quarter_bonus_refs)})" if quarter_bonus_refs else "0"])
    ws[f"B{current_row}"].number_format = "0.00"
    for cell in ws[current_row]:
        cell.font = Font(bold=True)
        cell.border = border
    current_row += 1

    # Special bonus hours
    ws.append(["Bonusberechtigte Stunden Sonderprojekt (Quartal):", f"=SUM({','.join(quarter_special_refs)})" if quarter_special_refs else "0"])
    ws[f"B{current_row}"].number_format = "0.00"
    for cell in ws[current_row]:
        cell.font = Font(bold=True)
        cell.border = border
    current_row += 1

    ws.append([])
    current_row += 1

    # Employee list
    ws.append(["--- Mitarbeiter in diesem Quartal ---"])
    ws[f"A{current_row}"].font = Font(bold=True, size=12)
    current_row += 1

    for emp in sorted(employee_summary_data.keys()):
        ws.append([emp])
        ws[f"A{current_row}"].border = border
        current_row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 35


def _add_vba_macro(xlsx_path: Path, progress_cb: ProgressCallback) -> Path:
    """
    Deprecated: COM automation removed.
    We now use a template with embedded VBA.
    """
    return xlsx_path


def build_quarterly_report(
    df_csv: pd.DataFrame,
    df_budget: pd.DataFrame,
    milestone_parent_map: Optional[Dict[Tuple[str, str], Set[str]]],
    df_xml: pd.DataFrame,
    target_quarter: pd.Period,
    months: Iterable[pd.Period],
    out_path: Path,
    progress_cb: ProgressCallback = _noop_progress,
    report_title: Optional[str] = None,
    use_quarter_filter: bool = True,
    add_vba: bool = True,
) -> Path:
    """Erstellt Quartals-Excel mit Monats-Tabellen + Quartalsübersicht."""

    if use_quarter_filter:
        df_quarter = df_xml[df_xml["quarter"] == target_quarter].copy()
    else:
        df_quarter = df_xml.copy()



    # Load template if available
    template_path = Path(__file__).parent / "template.xlsm"
    used_template = False
    
    if template_path.exists():
        try:
            wb = load_workbook(template_path, keep_vba=True)
            used_template = True
            progress_cb(2, "Template geladen")
        except Exception as e:
            progress_cb(2, f"Fehler beim Laden des Templates: {e}. Erstelle leeres Workbook.")
            wb = Workbook()
            wb.remove(wb.active)
    else:
        progress_cb(2, "Kein Template gefunden. Erstelle leeres Workbook.")
        wb = Workbook()
        wb.remove(wb.active)

    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Create Projekt-Budget-Übersicht sheet first
    progress_cb(18, "Erstelle Projekt-Budget-Übersicht")
    _create_project_budget_sheet(wb, df_budget, border)

    # Build lookup for each Obermeilenstein (no aggregation across a project)
    budget_lookup: Dict[Tuple[str, str], Dict[str, float]] = {}
    lookup_id_to_budget: Dict[int, Dict[str, float]] = {}

    def _proj_keys_for_lookup(name: str) -> List[str]:
        if name is None:
            return []
        base = str(name).strip()
        if not base:
            return []
        keys = [base]
        first = base.split(maxsplit=1)[0].strip() if base.split() else base
        if first and first not in keys:
            keys.append(first)
        return keys

    lookup_id_map: Dict[Tuple[str, str], int] = {}

    for _, row in df_budget.iterrows():
        projekt = str(row.get("Projekt", "")).strip()
        ober_norm = norm_ms(row.get("Obermeilenstein"))
        if not projekt or projekt in ("-", "nan") or not ober_norm:
            continue
        budget_data = {
            "Sollhonorar": float(row.get("Gesamtbudget") or 0.0),
            "Verrechnete_Honorare": float(row.get("Abgerechnet") or 0.0),
            "Istkosten": float(row.get("Istkosten") or 0.0),
            "Abrechnungsart": str(row.get("Abrechnungsart") or "").strip(),
            "LookupId": int(row.get("_LookupId") or 0),
            "Obermeilenstein_norm": ober_norm,
            "Sollstunden": float(row.get("Sollstunden") or 0.0),
        }
        for key in _proj_keys_for_lookup(projekt):
            budget_lookup[(key, ober_norm)] = budget_data
        if budget_data["LookupId"]:
            lookup_id_to_budget[budget_data["LookupId"]] = budget_data
            if budget_data["LookupId"]:
                lookup_id_map[(key, ober_norm)] = budget_data["LookupId"]

    def _resolve_budget_data(proj_value: str, ms_value: str) -> Optional[Dict[str, float]]:
        """Finds budget data for a (project, milestone) combination."""
        if not proj_value:
            return None
        proj_variants = _proj_keys_for_lookup(proj_value)
        ms_norm_value = norm_ms(ms_value)
        if not ms_norm_value:
            return None
        candidates: List[Tuple[str, str]] = []
        for key in proj_variants:
            candidates.append((key, ms_norm_value))
            if milestone_parent_map:
                parents = milestone_parent_map.get((key, ms_norm_value))
                if parents and len(parents) == 1:
                    parent_value = next(iter(parents))
                    candidates.append((key, parent_value))
        for candidate in candidates:
            if candidate in budget_lookup:
                return budget_lookup[candidate]
        return None

    def _build_lookup_expr(col_letter: str, primary_id: Optional[int], fallback_id: Optional[int]) -> str:
        def _index_expr(lookup_id: int) -> str:
            return (
                f"INDEX('Projekt-Budget-Übersicht'!${col_letter}:${col_letter},"
                f"MATCH({lookup_id},'Projekt-Budget-Übersicht'!$L:$L,0))"
            )

        def _safe_expr(lookup_id: int) -> str:
            return f"IFERROR({_index_expr(lookup_id)},0)"

        if primary_id is None:
            if fallback_id is not None:
                return _safe_expr(fallback_id)
            return "0"

        primary_expr = _safe_expr(primary_id)
        if fallback_id is not None:
            fallback_expr = _safe_expr(fallback_id)
            return f"IF({primary_expr}=0,{fallback_expr},{primary_expr})"
        return primary_expr

    def _determine_lookup_ids(proj_value: str, ms_value: str) -> Tuple[Optional[int], Optional[int]]:
        if not proj_value or not ms_value:
            return None, None

        ms_norm_value = norm_ms(ms_value)
        proj_variants = _proj_keys_for_lookup(proj_value)

        primary_id = None
        for key in proj_variants:
            candidate = (key, ms_norm_value)
            lookup_id = lookup_id_map.get(candidate)
            if lookup_id:
                primary_id = lookup_id
                break

        fallback_id = None
        if primary_id is None and milestone_parent_map:
            parent_norm = None
            for key in proj_variants:
                parents = milestone_parent_map.get((key, ms_norm_value))
                if parents:
                    parent_norm = sorted(parents)[0]
                    break
            if parent_norm:
                for key in proj_variants:
                    lookup_id = lookup_id_map.get((key, parent_norm))
                    if lookup_id:
                        fallback_id = lookup_id
                        break

        return primary_id, fallback_id

    employees = sorted(df_quarter["staff_name"].unique())
    total_emps = max(len(employees), 1)

    # Build a map of which employees work on which project/milestone combinations PER MONTH
    # Format: {(proj_norm, ms_norm, month): [list of employee names]}
    project_milestone_employees = {}
    for _, row in df_quarter.iterrows():
        key = (row["proj_norm"], row["ms_norm"], row["period"])
        emp_name = row["staff_name"]
        if key not in project_milestone_employees:
            project_milestone_employees[key] = set()
        project_milestone_employees[key].add(emp_name)
    # Convert sets to sorted lists
    for key in project_milestone_employees:
        project_milestone_employees[key] = sorted(list(project_milestone_employees[key]))

    # Dictionary to store cell references for summary sheet
    employee_summary_data = {}

    # Dictionary to track row assignments for "Von anderen" formula generation
    # Format: {employee: {(proj_norm, ms_norm, month): row_number}}
    row_assignments = {}

    # Dictionary to track month sections for summing "Von anderen" cells
    # Format: {employee: {month: (start_row, end_row, assigned_from_others_cell_row)}}
    month_sections = {}

    revenue_cells_by_key: Dict[Tuple[str, str, pd.Period], List[Tuple[str, int]]] = {}

    # Track quarterly row assignments across all employees
    # Format: {employee: {(proj_norm, ms_norm): row_number}}
    quarter_row_assignments_all = {}

    # Track quarterly revenue cells across all employees
    # Format: {(proj_norm, ms_norm): [(sheet_name, row_number), ...]}
    revenue_cells_by_key_q_all: Dict[Tuple[str, str], List[Tuple[str, int]]] = {}

    for idx_emp, emp in enumerate(employees, start=1):
        row_assignments[emp] = {}
        month_sections[emp] = {}
        quarter_row_assignments_all[emp] = {}
        ws = wb.create_sheet(title=emp[:31])
        sheet_name = emp[:31]
        monthly_bonus_total_cells: List[str] = []
        monthly_special_bonus_total_cells: List[str] = []
        transfer_entries: List[Tuple[str, str, str, str]] = []

        # Store monthly data for summary sheet
        if emp not in employee_summary_data:
            employee_summary_data[emp] = {
                'sheet_name': sheet_name,
                'months': {}
            }

        ws.append([f"{emp} - Quartalsreport {target_quarter}"])

        # Position dropdown for employee
        ws.append(["Position:", "SV"])  # Default: SV
        position_cell = ws.cell(row=2, column=2)
        position_dv = DataValidation(type="list", formula1='"SV,CAD,ADM,Pauschale,-"', allow_blank=False)
        position_dv.add(position_cell)
        ws.add_data_validation(position_dv)
        ws.cell(row=2, column=1).font = Font(bold=True)

        ws.append([])

        current_row = 4
        total_hours_all_months = 0.0
        total_bonus_hours_quarter = 0.0
        total_bonus_special_hours_quarter = 0.0

        # Track monthly summary row coordinates for quarterly summaries
        monthly_sum_total_cells = []
        monthly_assigned_from_others_cells = []
        monthly_bonus_base_cells = []  # Track G cells for bonus basis
        monthly_special_base_cells = []  # Track G cells for special bonus basis

        for month in months:
            df_month = df_quarter[(df_quarter["period"] == month) & (df_quarter["staff_name"] == emp)].copy()

            if df_month.empty:
                continue

            month_hours = (
                df_month.groupby(['proj_norm', 'ms_norm'], as_index=False)
                .agg({'hours': 'sum'})
            )

            month_data = month_hours.merge(
                df_csv,
                how="left",
                left_on=["proj_norm", "ms_norm"],
                right_on=["proj_norm", "ms_norm"],
            )

            month_data["Projekte"] = month_data["Projekte"].fillna(month_data["proj_norm"])
            month_data["Meilenstein"] = month_data["Meilenstein"].fillna(month_data["ms_norm"])
            month_data["MeilensteinTyp"] = month_data["Meilenstein"].apply(get_milestone_type)

            month_data["Soll"] = month_data["Soll"].fillna(0.0)
            month_data["Ist"] = month_data["Ist"].fillna(0.0)

            for idx in month_data.index:
                if month_data.loc[idx, "Soll"] == 0.0 and month_data.loc[idx, "Ist"] == 0.0:
                    ms_name = month_data.loc[idx, "Meilenstein"]
                    ms_type = month_data.loc[idx, "MeilensteinTyp"]

                    if ms_type == "monthly" and ms_name in MONTHLY_BUDGETS:
                        # Monthly budget (NOT cumulative - each month has its own budget)
                        month_data.loc[idx, "Soll"] = MONTHLY_BUDGETS[ms_name]
                        month_data.loc[idx, "Ist"] = month_data.loc[idx, "hours"]

            for idx in month_data.index:
                ms_name = month_data.loc[idx, "Meilenstein"]
                ms_type = month_data.loc[idx, "MeilensteinTyp"]
                proj_name = month_data.loc[idx, "Projekte"]
                proj_norm = month_data.loc[idx, "proj_norm"] if "proj_norm" in month_data.columns else ""
                if ms_type == "monthly" and ms_name in MONTHLY_BUDGETS and (is_bonus_project(proj_name) or is_bonus_project(proj_norm)):
                    # Monthly budget (NOT cumulative - each month has its own budget)
                    month_data.loc[idx, "Soll"] = MONTHLY_BUDGETS[ms_name]
                    month_data.loc[idx, "Ist"] = month_data.loc[idx, "hours"]

            def _compute_month_qsoll(row):
                if row.get("MeilensteinTyp") != "quarterly":
                    return 0.0
                hours, unit = extract_budget_from_name(row.get("Meilenstein"))
                if unit == "quartal" and hours is not None:
                    return hours
                name = row.get("Meilenstein")
                if name in QUARTERLY_BUDGETS:
                    return QUARTERLY_BUDGETS[name]
                try:
                    if pd.notna(row.get("Soll")) and float(row.get("Soll")) > 0:
                        return float(row.get("Soll"))
                except Exception:
                    pass
                return 0.0

            month_data["QuartalsSoll"] = month_data.apply(_compute_month_qsoll, axis=1)

            # Cumulative XML hours up to current month (for quarterly milestones)
            df_to_date = df_quarter[(df_quarter["staff_name"] == emp) & (df_quarter["period"] <= month)]
            cum_hours_map = {
                (r["proj_norm"], r["ms_norm"]): r["hours"]
                for _, r in df_to_date.groupby(["proj_norm", "ms_norm"], as_index=False).agg({"hours": "sum"}).iterrows()
            }

            # XML hours for months AFTER the current month - FOR ALL EMPLOYEES (for backward calculation)
            # This ensures all employees see the same IST value for the same project/milestone
            df_after_month_all_employees = df_quarter[df_quarter["period"] > month]
            future_hours_all_employees_map = {
                (r["proj_norm"], r["ms_norm"]): r["hours"]
                for _, r in df_after_month_all_employees.groupby(["proj_norm", "ms_norm"], as_index=False).agg({"hours": "sum"}).iterrows()
            }

            month_data = month_data.sort_values(["Projekte", "Meilenstein"])

            month_name = MONTH_NAMES.get(int(month.month), month.strftime('%B'))
            month_str = f"{month_name} {month.year}"

            ws.append([f"--- {month_str} ---"])
            ws[f"A{current_row}"].font = Font(bold=True, size=12)
            current_row += 1

            ws.append(["Projekt", "Meilenstein", "Abrechnungsart", "Soll (h)", "Ist (h)", f"{month_str} (h)", "%", "Bonus-Anpassung (h)", "Differenz (h)", "Zuordnen an", "Von anderen (h)", "Stundensatz (€/h)", "Umsatz (€)", "Möglicher Umsatz (€)", "Entgangener Umsatz (€)", "Umsatz kumuliert (€)", "Soll Obermeilenstein (h)", "Budget Gesamt (€)", "Kosten (€)", "Rechnung", "Kommentar"])
            for cell in ws[current_row]:
                cell.font = Font(bold=True)
                cell.border = border
            current_row += 1

            # Track start of month data section
            month_data_start_row = current_row

            bonus_hours_month = 0.0
            bonus_hours_month_special = 0.0
            adjustment_cells_regular: List[str] = []
            adjustment_cells_special: List[str] = []

            for proj, proj_block in month_data.groupby("Projekte", sort=False):
                proj_block = proj_block.reset_index(drop=True)
                block_start = current_row

                for i, (_, row_data) in enumerate(proj_block.iterrows()):
                    ms_type = row_data["MeilensteinTyp"]
                    hours_value = float(row_data.get("hours") or 0.0)
                    is_special_project = is_bonus_project(proj) or is_bonus_project(row_data.get("proj_norm", ""))

                    # Get billing type from budget data
                    projekt_name = row_data["proj_norm"]
                    meilenstein_name = row_data["ms_norm"]
                    resolved_budget = _resolve_budget_data(projekt_name, meilenstein_name)
                    billing_type_display = (resolved_budget.get("Abrechnungsart", "").strip()
                                           if resolved_budget else "Unbekannt")
                    bonus_candidate = False
                    should_color = False
                    color_percentage = 0.0

                    if ms_type == "monthly":
                        csv_ist_total = float(row_data.get("Ist") or 0.0)
                        ms_name = row_data["Meilenstein"]

                        # For 0000-projects: use per-employee budget, not cumulative
                        if is_special_project:
                            # Check MONTHLY_BUDGETS first
                            if ms_name in MONTHLY_BUDGETS:
                                soll_value = MONTHLY_BUDGETS[ms_name]
                                ist_display = hours_value
                            else:
                                # Try to extract from name
                                hours, unit = extract_budget_from_name(ms_name)
                                if unit == "monat" and hours is not None:
                                    soll_value = hours
                                    ist_display = hours_value
                                else:
                                    soll_value = float(row_data.get("Soll") or 0.0)
                                    ist_display = hours_value
                        else:
                            soll_value = float(row_data.get("Soll") or 0.0)
                            # For normal projects: backward calculation from CSV IST
                            # Using ALL EMPLOYEES' future hours to ensure consistent IST across all employees
                            # Last month: IST = CSV_IST
                            # Previous months: IST = CSV_IST - XML_hours_of_future_months_ALL_EMPLOYEES
                            key = (row_data["proj_norm"], row_data["ms_norm"])
                            xml_future_hours_all = float(future_hours_all_employees_map.get(key, 0.0))
                            ist_display = csv_ist_total - xml_future_hours_all

                        if soll_value > 0:
                            pct_value = (ist_display / soll_value) * 100.0 if soll_value else 0.0
                            should_color = True
                            color_percentage = pct_value
                            if pct_value <= 100.0:
                                bonus_candidate = True
                        else:
                            pct_value = 0.0
                            bonus_candidate = True
                        ws.append([
                            proj if i == 0 else "",
                            row_data["Meilenstein"],
                            None,  # Abrechnungsart (C) - Formula will be added later
                            round(soll_value, 2),
                            round(ist_display, 2),
                            round(hours_value, 2),
                            round(pct_value, 2),
                            None,  # Bonus-Anpassung (H)
                            None,  # Differenz (I) - will be filled with formula
                            None,  # Zuordnen an (J) - Dropdown will be added later
                            None,  # Von anderen (K) - will be filled with formula
                            None,  # Stundensatz (L) - Formula will be added later
                            None,  # Umsatz (M) - Formula will be added later
                            None,  # Möglicher Umsatz (N) - Formula will be added later
                            None,  # Entgangener Umsatz (O) - Formula will be added later: =N-M
                            None,  # Umsatz kumuliert (P) - Formula will be added later
                            None,  # Soll Obermeilenstein (Q) - Formula will be added later
                            None,  # Budget Gesamt (R) - Formula will be added later
                            None,  # Kosten (S) - From CSV
                            None,  # Rechnung (T) - Dropdown will be added later
                            None,  # Kommentar (U) - Empty field for user input
                        ])
                    else:
                        q_soll = float(row_data.get("QuartalsSoll", 0.0) or 0.0)
                        cum_ist = float(cum_hours_map.get((row_data["proj_norm"], row_data["ms_norm"]), 0.0))
                        prozent = (cum_ist / q_soll * 100.0) if q_soll > 0 else 0.0
                        if prozent <= 100.0:
                            bonus_candidate = True
                        should_color = q_soll > 0
                        color_percentage = prozent
                        ws.append([
                            proj if i == 0 else "",
                            row_data["Meilenstein"],
                            None,  # Abrechnungsart (C) - Formula will be added later
                            round(q_soll, 2) if q_soll > 0 else "-",
                            round(cum_ist, 2) if cum_ist > 0 else 0.0,
                            round(hours_value, 2),
                            round(prozent, 2) if q_soll > 0 else "-",
                            None,  # Bonus-Anpassung (H)
                            None,  # Differenz (I) - will be filled with formula
                            None,  # Zuordnen an (J) - Dropdown will be added later
                            None,  # Von anderen (K) - will be filled with formula
                            None,  # Stundensatz (L) - Formula will be added later
                            None,  # Umsatz (M) - Formula will be added later
                            None,  # Budget Gesamt (N) - Formula will be added later
                            None,  # Kosten (O) - From CSV
                            None,  # Umsatz kumuliert (P) - Formula will be added later
                        ])

                    for cell in ws[current_row]:
                        cell.border = border

                    # Bonus-Anpassung cell (column H)
                    adj_cell = ws.cell(row=current_row, column=8)
                    if is_special_project:
                        adjustment_cells_special.append(adj_cell.coordinate)
                    else:
                        adjustment_cells_regular.append(adj_cell.coordinate)
                    adj_cell.number_format = "0.00"

                    # Differenz cell (column I) - use negative adjustment as transfer amount, never below 0
                    diff_cell = ws.cell(row=current_row, column=9)
                    diff_cell.value = f"=IF(H{current_row}<0,MAX(0,MIN(F{current_row},-H{current_row})),0)"
                    diff_cell.number_format = "0.00"

                    # Zuordnen an cell (column J) - Dropdown with other employees on same project/milestone IN SAME MONTH
                    assign_cell = ws.cell(row=current_row, column=10)
                    key = (row_data["proj_norm"], row_data["ms_norm"], month)
                    other_employees = [e for e in project_milestone_employees.get(key, []) if e != emp]
                    if other_employees:
                        # Create dropdown with other employees
                        employee_list = ",".join(other_employees)
                        dv = DataValidation(type="list", formula1=f'"{employee_list}"', allow_blank=True)
                        dv.add(assign_cell)
                        ws.add_data_validation(dv)

                    # Add conditional formatting to turn cell red when negative
                    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    red_font = Font(color='9C0006')
                    ws.conditional_formatting.add(
                        diff_cell.coordinate,
                        CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font)
                    )

                    # Von anderen cell (column K) - Formula to sum hours assigned by other employees
                    # This will be filled in a second pass after all sheets are created
                    from_others_cell = ws.cell(row=current_row, column=11)
                    from_others_cell.number_format = "0.00"
                    # Mark this cell with row info for later formula injection
                    from_others_cell.value = 0  # Placeholder, will be replaced with formula

                    proj_norm_value = row_data["proj_norm"]
                    ms_norm_value = row_data["ms_norm"]
                    lookup_primary_id, lookup_fallback_id = _determine_lookup_ids(proj_norm_value, ms_norm_value)

                    # Abrechnungsart cell (column C) - Lookup from Projekt-Budget-Übersicht
                    billing_cell = ws.cell(row=current_row, column=3)
                    if lookup_primary_id or lookup_fallback_id:
                        lookup_id = lookup_primary_id if lookup_primary_id else lookup_fallback_id
                        billing_formula = (
                            f'=IFERROR(INDEX(\'Projekt-Budget-Übersicht\'!$C:$C,'
                            f'MATCH({lookup_id},\'Projekt-Budget-Übersicht\'!$L:$L,0)),"")'
                        )
                        billing_cell.value = billing_formula
                    else:
                        billing_cell.value = "Unbekannt"

                    # Stundensatz cell (column L) - Lookup via Schlüssel
                    rate_cell = ws.cell(row=current_row, column=12)
                    rate_cell.number_format = '#,##0.00'

                    # Build simplified formula with single IFERROR wrapper
                    if lookup_primary_id or lookup_fallback_id:
                        lookup_id = lookup_primary_id if lookup_primary_id else lookup_fallback_id
                        rate_formula = (
                            f'=IF($B$2="-",0,'
                            f'IFERROR(INDEX('
                            f'IF($B$2="SV",\'Projekt-Budget-Übersicht\'!$I:$I,'
                            f'IF($B$2="CAD",\'Projekt-Budget-Übersicht\'!$J:$J,'
                            f'IF($B$2="ADM",\'Projekt-Budget-Übersicht\'!$K:$K,'
                            f'\'Projekt-Budget-Übersicht\'!$I:$I))),'  # Default to SV/Pauschale
                            f'MATCH({lookup_id},\'Projekt-Budget-Übersicht\'!$M:$M,0)),0))'
                        )
                    else:
                        rate_formula = "=IF($B$2=\"-\",0,0)"
                    rate_cell.value = rate_formula

                    projekt_name = row_data["proj_norm"]
                    meilenstein_name = row_data["ms_norm"]
                    resolved_budget = _resolve_budget_data(projekt_name, meilenstein_name)
                    billing_type = (resolved_budget.get("Abrechnungsart", "").strip()
                                    if resolved_budget else "")

                    # Umsatz cell (column M) - Formula reads Abrechnungsart from column C dynamically
                    # Rules: 1) No adjustments (I) in revenue calculation
                    #        2) Pauschale: Revenue capped at remaining budget up to 100%
                    #        3) Nachweis: ALWAYS full amount (hourly rate × hours worked)
                    revenue_cell = ws.cell(row=current_row, column=13)
                    revenue_cell.number_format = '#,##0.00'
                    # Dynamic formula that checks column C (Abrechnungsart)
                    # For Pauschale: MIN(monthly proportional revenue, remaining budget until 100%)
                    revenue_formula = (
                        f'=IF($B$2="-",0,'
                        f'IF(C{current_row}="Pauschale",'
                        f'IF(OR(R{current_row}=0,N(D{current_row})=0),0,'
                        f'MAX(0,MIN(R{current_row}*((N(F{current_row})+N(K{current_row}))/N(D{current_row})),'
                        f'R{current_row}-R{current_row}*(N(E{current_row})/N(D{current_row}))))),'
                        f'L{current_row}*(N(F{current_row})+N(K{current_row}))))'
                    )
                    revenue_cell.value = revenue_formula

                    # Möglicher Umsatz cell (column N) - Same formula but WITHOUT >100% check
                    # Shows what COULD be billed (without >100% restriction for Pauschale)
                    possible_revenue_cell = ws.cell(row=current_row, column=14)
                    possible_revenue_cell.number_format = '#,##0.00'
                    possible_revenue_formula = (
                        f'=IF($B$2="-",0,'
                        f'IF(C{current_row}="Pauschale",'
                        f'IF(OR(R{current_row}=0,N(D{current_row})=0),0,'
                        f'R{current_row}*((N(F{current_row})+N(K{current_row}))/N(D{current_row}))),'
                        f'L{current_row}*(N(F{current_row})+N(K{current_row}))))'
                    )
                    possible_revenue_cell.value = possible_revenue_formula

                    # Entgangener Umsatz cell (column O) - Difference between Möglicher Umsatz and Umsatz
                    lost_revenue_cell = ws.cell(row=current_row, column=15)
                    lost_revenue_cell.number_format = '#,##0.00'
                    lost_revenue_cell.value = f"=N{current_row}-M{current_row}"

                    # Umsatz kumuliert cell (column P) - Placeholder, will be filled in second pass
                    budget_earned_cell = ws.cell(row=current_row, column=16)
                    budget_earned_cell.number_format = '#,##0.00'
                    budget_earned_cell.value = 0

                    # Soll Obermeilenstein cell (column Q) - direct value from resolved budget (fallback to parent)
                    soll_obermeilenstein_cell = ws.cell(row=current_row, column=17)
                    soll_obermeilenstein_cell.number_format = '#,##0.00'
                    budget_record = resolved_budget
                    if not budget_record and lookup_primary_id:
                        budget_record = lookup_id_to_budget.get(lookup_primary_id)
                    if not budget_record and lookup_fallback_id:
                        budget_record = lookup_id_to_budget.get(lookup_fallback_id)
                    if budget_record:
                        soll_val = float(budget_record.get("Sollstunden") or 0.0)
                        soll_obermeilenstein_cell.value = soll_val if soll_val != 0 else ""
                    else:
                        soll_obermeilenstein_cell.value = ""

                    # Budget Gesamt cell (column R) - Lookup via Schlüssel
                    budget_total_cell = ws.cell(row=current_row, column=18)
                    budget_total_cell.number_format = '#,##0.00'
                    budget_expr = _build_lookup_expr("F", lookup_primary_id, lookup_fallback_id)
                    budget_total_cell.value = f"={budget_expr}"

                    # Kosten cell (column S) - Real costs (Istkosten) from CSV
                    budget_ist_cell = ws.cell(row=current_row, column=19)
                    budget_ist_cell.number_format = '#,##0.00'
                    if resolved_budget:
                        istkosten_value = resolved_budget.get("Istkosten", 0) or 0
                        budget_ist_cell.value = istkosten_value if istkosten_value != 0 else ""
                    else:
                        budget_ist_cell.value = ""

                    # Rechnung cell (column T) - Dropdown with SR/AZ options
                    rechnung_cell = ws.cell(row=current_row, column=20)
                    rechnung_dv = DataValidation(type="list", formula1='"SR,AZ"', allow_blank=True)
                    rechnung_dv.add(rechnung_cell)
                    ws.add_data_validation(rechnung_dv)

                    # Kommentar cell (column U) - Empty field for user input
                    kommentar_cell = ws.cell(row=current_row, column=21)
                    # Leave empty for user input

                    # Track row for this project/milestone/month combination
                    track_key = (row_data["proj_norm"], row_data["ms_norm"], month)
                    row_assignments[emp][track_key] = current_row
                    rev_ref = f"'{sheet_name}'!M{current_row}"
                    revenue_cells_by_key.setdefault(track_key, []).append((sheet_name, current_row))

                    if should_color:
                        pct_cell = ws.cell(row=current_row, column=7)
                        pct_cell.fill = PatternFill(
                            start_color=status_color_hex(color_percentage),
                            end_color=status_color_hex(color_percentage),
                            fill_type="solid",
                        )

                    if bonus_candidate:
                        if is_special_project:
                            bonus_hours_month_special += hours_value
                        else:
                            bonus_hours_month += hours_value

                    current_row += 1

                block_size = len(proj_block)
                if block_size > 1:
                    ws.merge_cells(start_row=block_start, start_column=1,
                                   end_row=block_start + block_size - 1, end_column=1)
                    ws.cell(row=block_start, column=1).alignment = Alignment(vertical="top")

            # Track end of month data section (before summary rows)
            month_data_end_row = current_row - 1

            sum_hours = month_data["hours"].sum()
            total_hours_all_months += sum_hours
            ws.append(["", "Summe", "", "", "", round(sum_hours, 2), "", "", "", "", "", "", "", "", "", ""])
            sum_row_idx = current_row
            for cell in ws[current_row]:
                cell.font = Font(bold=True)
                cell.border = border
            sum_total_cell = ws.cell(row=sum_row_idx, column=6)
            sum_total_cell.number_format = "0.00"
            sum_total_cell.value = round(sum_hours, 2)

            # Add sum formulas for Umsatz (M), Möglicher Umsatz (N), Entgangener Umsatz (O), and Umsatz kumuliert (P)
            sum_umsatz_cell = ws.cell(row=sum_row_idx, column=13)  # Column M
            sum_umsatz_cell.number_format = '#,##0.00'
            if month_data_start_row <= month_data_end_row:
                sum_umsatz_cell.value = f"=SUM(M{month_data_start_row}:M{month_data_end_row})"
            else:
                sum_umsatz_cell.value = 0

            sum_possible_umsatz_cell = ws.cell(row=sum_row_idx, column=14)  # Column N
            sum_possible_umsatz_cell.number_format = '#,##0.00'
            if month_data_start_row <= month_data_end_row:
                sum_possible_umsatz_cell.value = f"=SUM(N{month_data_start_row}:N{month_data_end_row})"
            else:
                sum_possible_umsatz_cell.value = 0

            sum_lost_umsatz_cell = ws.cell(row=sum_row_idx, column=15)  # Column O
            sum_lost_umsatz_cell.number_format = '#,##0.00'
            if month_data_start_row <= month_data_end_row:
                sum_lost_umsatz_cell.value = f"=SUM(O{month_data_start_row}:O{month_data_end_row})"
            else:
                sum_lost_umsatz_cell.value = 0

            sum_umsatz_kum_cell = ws.cell(row=sum_row_idx, column=16)  # Column P
            sum_umsatz_kum_cell.number_format = '#,##0.00'
            if month_data_start_row <= month_data_end_row:
                sum_umsatz_kum_cell.value = f"=SUM(P{month_data_start_row}:P{month_data_end_row})"
            else:
                sum_umsatz_kum_cell.value = 0

            current_row += 1

            ws.append(["", "Bonusberechtigte Stunden", "", "", "", 0, "", "", "", "", "", "", "", "", "", ""])
            bonus_row_idx = current_row
            for cell in ws[current_row]:
                cell.font = Font(bold=True)
                cell.border = border
            bonus_base_cell = ws.cell(row=bonus_row_idx, column=7)
            bonus_base_cell.number_format = "0.00"
            bonus_base_cell.value = round(bonus_hours_month, 2)
            bonus_total_cell = ws.cell(row=bonus_row_idx, column=6)
            bonus_total_cell.number_format = "0.00"
            if adjustment_cells_regular:
                adj_formula = ",".join(adjustment_cells_regular)
                adj_sum_formula = f"SUM({adj_formula})"
                bonus_adj_cell = ws.cell(row=bonus_row_idx, column=8)
                bonus_adj_cell.value = f"={adj_sum_formula}"
                bonus_adj_cell.number_format = "0.00"
                bonus_total_cell.value = f"={bonus_base_cell.coordinate}+{bonus_adj_cell.coordinate}"
            else:
                bonus_total_cell.value = round(bonus_hours_month, 2)
                bonus_adj_cell = ws.cell(row=bonus_row_idx, column=8)
                bonus_adj_cell.value = 0
                bonus_adj_cell.number_format = "0.00"
            monthly_bonus_total_cells.append(bonus_total_cell.coordinate)
            monthly_sum_total_cells.append(sum_total_cell.coordinate)
            monthly_bonus_base_cells.append(bonus_base_cell.coordinate)
            total_bonus_hours_quarter += bonus_hours_month
            current_row += 1

            ws.append(["", "Bonusberechtigte Stunden Sonderprojekt", "", "", "", 0, "", "", "", "", "", "", "", "", "", ""])
            special_row_idx = current_row
            for cell in ws[current_row]:
                cell.font = Font(bold=True)
                cell.border = border
            special_base_cell = ws.cell(row=special_row_idx, column=7)
            special_base_cell.number_format = "0.00"
            special_base_cell.value = round(bonus_hours_month_special, 2)
            special_total_cell = ws.cell(row=special_row_idx, column=6)
            special_total_cell.number_format = "0.00"
            if adjustment_cells_special:
                adj_formula_special = ",".join(adjustment_cells_special)
                adj_sum_formula_special = f"SUM({adj_formula_special})"
                special_adj_cell = ws.cell(row=special_row_idx, column=8)
                special_adj_cell.value = f"={adj_sum_formula_special}"
                special_adj_cell.number_format = "0.00"
                special_total_cell.value = f"={special_base_cell.coordinate}+{special_adj_cell.coordinate}"
            else:
                special_total_cell.value = round(bonus_hours_month_special, 2)
                special_adj_cell = ws.cell(row=special_row_idx, column=8)
                special_adj_cell.value = 0
                special_adj_cell.number_format = "0.00"
            monthly_special_bonus_total_cells.append(special_total_cell.coordinate)
            monthly_special_base_cells.append(special_base_cell.coordinate)
            total_bonus_special_hours_quarter += bonus_hours_month_special
            current_row += 1

            # Zugeordnete Stunden von anderen MA - will be calculated with formula
            ws.append(["", "Zugeordnete Stunden von anderen MA", "", "", "", 0, "", "", "", "", "", "", "", "", "", ""])
            assigned_from_others_row_idx = current_row
            for cell in ws[current_row]:
                cell.font = Font(bold=True)
                cell.border = border
            assigned_from_others_cell = ws.cell(row=assigned_from_others_row_idx, column=6)
            assigned_from_others_cell.number_format = "0.00"
            # Formula will sum all "Von anderen (K)" cells in this month's section
            # This will be filled after all sheets are created
            assigned_from_others_cell.value = 0  # Placeholder
            monthly_assigned_from_others_cells.append(assigned_from_others_cell.coordinate)

            # Track month section info
            month_sections[emp][month] = (month_data_start_row, month_data_end_row, assigned_from_others_row_idx)

            current_row += 1

            # Gesamt Bonus Stunden = Bonusberechtigte + Sonderprojekt + Zugeordnete
            ws.append(["", "Gesamt Bonus Stunden", "", "", "", 0, "", "", "", "", ""])
            total_bonus_row_idx = current_row
            for cell in ws[current_row]:
                cell.font = Font(bold=True)
                cell.border = border
                cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
            total_bonus_cell = ws.cell(row=total_bonus_row_idx, column=6)
            total_bonus_cell.number_format = "0.00"
            total_bonus_cell.value = f"={bonus_total_cell.coordinate}+{special_total_cell.coordinate}+{assigned_from_others_cell.coordinate}"
            current_row += 1

            transfer_entries.append((month_str, sum_total_cell.coordinate, bonus_total_cell.coordinate, special_total_cell.coordinate, assigned_from_others_cell.coordinate, total_bonus_cell.coordinate))

            # Store cell references for summary sheet
            employee_summary_data[emp]['months'][month_str] = {
                'total_hours_cell': f"'{sheet_name}'!{sum_total_cell.coordinate}",
                'bonus_hours_cell': f"'{sheet_name}'!{bonus_total_cell.coordinate}",
                'special_bonus_hours_cell': f"'{sheet_name}'!{special_total_cell.coordinate}"
            }

            ws.append([])
            current_row += 1

        if transfer_entries:
            ws.append(["--- Übertragshilfe ---"])
            ws[f"A{current_row}"].font = Font(bold=True, size=12)
            current_row += 1

            ws.append(["Monat", "Mitarbeiter", "Prod. Stunden", "Bonusberechtigte Stunden", "Bonusberechtigte Stunden Sonderprojekt", "Zugeordnet von anderen", "Gesamt Bonus"])
            for cell in ws[current_row]:
                cell.font = Font(bold=True)
                cell.border = border
            current_row += 1

            for month_label, total_cell, bonus_cell, special_cell, assigned_cell, total_bonus_cell in transfer_entries:
                ws.append([month_label, emp, f"={total_cell}", f"={bonus_cell}", f"={special_cell}", f"={assigned_cell}", f"={total_bonus_cell}"])
                for cell in ws[current_row]:
                    cell.border = border
                current_row += 1

            ws.append([])
            current_row += 1

        df_emp_quarter = df_quarter[df_quarter["staff_name"] == emp].copy()

        if df_emp_quarter.empty:
            continue

        quarter_hours = (
            df_emp_quarter.groupby(['proj_norm', 'ms_norm'], as_index=False)
            .agg({'hours': 'sum'})
        )

        quarter_data = quarter_hours.merge(
            df_csv,
            how="left",
            left_on=["proj_norm", "ms_norm"],
            right_on=["proj_norm", "ms_norm"],
        )

        quarter_data["Projekte"] = quarter_data["Projekte"].fillna(quarter_data["proj_norm"])
        quarter_data["Meilenstein"] = quarter_data["Meilenstein"].fillna(quarter_data["ms_norm"])
        quarter_data["MeilensteinTyp"] = quarter_data["Meilenstein"].apply(get_milestone_type)

        quarter_quarterly = quarter_data[quarter_data["MeilensteinTyp"] == "quarterly"].copy()

        def _compute_quarter_soll(row):
            hours, unit = extract_budget_from_name(row.get("Meilenstein"))
            if unit == "quartal" and hours is not None:
                return hours
            if row.get("Meilenstein") in QUARTERLY_BUDGETS:
                return QUARTERLY_BUDGETS[row.get("Meilenstein")]
            soll_val = row.get("Soll", np.nan)
            try:
                if pd.notna(soll_val) and float(soll_val) > 0:
                    return float(soll_val)
            except Exception:
                pass
            return 0.0

        quarter_quarterly["QuartalsSoll"] = quarter_quarterly.apply(_compute_quarter_soll, axis=1)

        if not quarter_quarterly.empty:
            ws.append([f"--- Quartalsübersicht {target_quarter} ---"])
            ws[f"A{current_row}"].font = Font(bold=True, size=12)
            current_row += 1

            ws.append(["Projekt", "Meilenstein", "Q-Soll (h)", "Q-Ist (h)", "%"])
            for cell in ws[current_row]:
                cell.font = Font(bold=True)
                cell.border = border
            current_row += 1

            for proj, proj_block in quarter_quarterly.groupby("Projekte", sort=False):
                proj_block = proj_block.reset_index(drop=True)
                block_start = current_row

                for i, (_, row_data) in enumerate(proj_block.iterrows()):
                    ms_name = row_data["Meilenstein"]
                    q_soll = row_data.get("QuartalsSoll", 0.0)
                    q_ist = row_data["hours"]
                    prozent = (q_ist / q_soll * 100.0) if q_soll > 0 else 0.0

                    ws.append([
                        proj if i == 0 else "",
                        ms_name,
                        round(q_soll, 2) if q_soll > 0 else "-",
                        round(q_ist, 2),
                        round(prozent, 2) if q_soll > 0 else "-"
                    ])

                    for cell in ws[current_row]:
                        cell.border = border

                    if q_soll > 0:
                        pct_cell = ws.cell(row=current_row, column=5)
                        pct_cell.fill = PatternFill(
                            start_color=status_color_hex(prozent),
                            end_color=status_color_hex(prozent),
                            fill_type="solid",
                        )
                    current_row += 1

                block_size = len(proj_block)
                if block_size > 1:
                    ws.merge_cells(start_row=block_start, start_column=1,
                                   end_row=block_start + block_size - 1, end_column=1)
                    ws.cell(row=block_start, column=1).alignment = Alignment(vertical="top")

        # ========== QUARTERLY SUMMARY TABLE ==========
        ws.append([])
        current_row += 1
        ws.append([f"--- Quartalszusammenfassung {target_quarter} ---"])
        ws[f"A{current_row}"].font = Font(bold=True, size=14)
        current_row += 1

        # Aggregate quarter data for this employee - sum hours across all months
        df_emp_quarter = df_quarter[df_quarter["staff_name"] == emp].copy()

        # Group by project and milestone across all months to get quarterly totals
        quarter_agg = (
            df_emp_quarter.groupby(['proj_norm', 'ms_norm'], as_index=False)
            .agg({'hours': 'sum'})
        )

        # Merge with CSV data to get project names and Soll values
        quarter_with_csv = pd.merge(
            quarter_agg,
            df_csv[['proj_norm', 'ms_norm', 'Projekte', 'Meilenstein', 'Soll', 'Ist']],
            on=['proj_norm', 'ms_norm'],
            how='left'
        )

        # Drop duplicates that might arise from merge
        quarter_with_csv = quarter_with_csv.drop_duplicates(subset=['proj_norm', 'ms_norm'])

        # Header row for quarterly table
        ws.append(["Projekt", "Meilenstein", "Abrechnungsart", "Soll (h)", "Ist (h)", "Quartal (h)", "%", "Bonus-Anpassung (h)", "Differenz (h)", "Zuordnen an", "Von anderen (h)", "Stundensatz (€/h)", "Umsatz (€)", "Möglicher Umsatz (€)", "Entgangener Umsatz (€)", "Umsatz kumuliert (€)", "Budget Gesamt (€)", "Kosten (€)"])
        for cell in ws[current_row]:
            cell.font = Font(bold=True)
            cell.border = border
        current_row += 1

        quarter_data_start_row = current_row
        adjustment_cells_regular_q = []
        adjustment_cells_special_q = []

        # Track F-column rows for regular and special projects (for base calculation)
        regular_project_rows = []
        special_project_rows = []

        # Track row assignments for quarterly "Von anderen" formulas
        # Format: {(proj_norm, ms_norm): row_number}
        quarter_row_assignments = {}

        # Track quarterly revenue cells by key
        revenue_cells_by_key_q: Dict[Tuple[str, str], List[Tuple[str, int]]] = {}

        # Process each project/milestone in the quarter
        for proj, proj_block in quarter_with_csv.groupby("Projekte", sort=False):
            proj_block = proj_block.reset_index(drop=True)
            block_start = current_row

            for i, (_, row_data) in enumerate(proj_block.iterrows()):
                hours_value = float(row_data.get("hours") or 0.0)
                is_special_project = is_bonus_project(proj) or is_bonus_project(row_data.get("proj_norm", ""))

                # Get budget data
                projekt_name = row_data["proj_norm"]
                meilenstein_name = row_data["ms_norm"]
                resolved_budget = _resolve_budget_data(projekt_name, meilenstein_name)

                # For 0000 projects, calculate Soll per employee (not cumulative)
                if is_special_project:
                    milestone_full_name = row_data.get("Meilenstein", "")
                    # Check if it's a quarterly or monthly 0000 milestone
                    hours, unit = extract_budget_from_name(milestone_full_name)
                    if unit == "quartal" and hours is not None:
                        q_soll = hours
                    elif unit == "monat" and hours is not None:
                        q_soll = hours * 3  # Monthly budget * 3 months
                    elif milestone_full_name in QUARTERLY_BUDGETS:
                        q_soll = QUARTERLY_BUDGETS[milestone_full_name]
                    elif milestone_full_name in MONTHLY_BUDGETS:
                        q_soll = MONTHLY_BUDGETS[milestone_full_name] * 3
                    else:
                        q_soll = 0.0
                    # Ist is the sum of hours from XML for this employee
                    ist_value = hours_value
                else:
                    q_soll = float(row_data.get("Soll", 0.0) or 0.0)
                    ist_value = float(row_data.get("Ist", 0.0) or 0.0)

                prozent = (ist_value / q_soll * 100.0) if q_soll > 0 else 0.0

                should_color = q_soll > 0

                # Append row
                ws.append([
                    proj if i == 0 else "",
                    row_data["Meilenstein"],
                    None,  # Abrechnungsart (C) - Formula will be added later
                    round(q_soll, 2) if q_soll > 0 else "-",
                    round(ist_value, 2) if ist_value > 0 else 0.0,
                    None,  # Quartal (h) (F) - Will be calculated as sum of monthly F+I+K
                    round(prozent, 2) if q_soll > 0 else "-",
                    None,  # Bonus-Anpassung (H)
                    None,  # Differenz (I)
                    "-",  # Zuordnen an (J) - Not applicable for quarterly view
                    None,  # Von anderen (K) - Will be calculated
                    None,  # Stundensatz (L)
                    None,  # Umsatz (M)
                    None,  # Budget Gesamt (N)
                    None,  # Kosten (O)
                    None,  # Umsatz kumuliert (P)
                ])

                for cell in ws[current_row]:
                    cell.border = border

                # Bonus-Anpassung cell (column H) - Sum of monthly adjustments for this project/milestone
                adj_cell = ws.cell(row=current_row, column=8)
                if is_special_project:
                    adjustment_cells_special_q.append(adj_cell.coordinate)
                else:
                    adjustment_cells_regular_q.append(adj_cell.coordinate)
                adj_cell.number_format = "0.00"

                # Collect monthly H values for this project/milestone to sum them
                monthly_adj_refs = []
                for month in months:
                    monthly_key = (projekt_name, meilenstein_name, month)
                    if monthly_key in row_assignments[emp]:
                        monthly_row = row_assignments[emp][monthly_key]
                        monthly_adj_refs.append(f"H{monthly_row}")

                # Set formula to sum monthly adjustments
                if monthly_adj_refs:
                    adj_cell.value = f"=SUM({','.join(monthly_adj_refs)})"
                else:
                    adj_cell.value = 0

                # Differenz cell (column I) - use negative adjustment as transfer amount, never below 0
                diff_cell = ws.cell(row=current_row, column=9)
                diff_cell.value = f"=IF(H{current_row}<0,MAX(0,MIN(F{current_row},-H{current_row})),0)"
                diff_cell.number_format = "0.00"

                # Add red conditional formatting
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                red_font = Font(color='9C0006')
                ws.conditional_formatting.add(
                    diff_cell.coordinate,
                    CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font)
                )

                # Von anderen cell (column K) - Placeholder
                from_others_cell_q = ws.cell(row=current_row, column=11)
                from_others_cell_q.number_format = "0.00"
                from_others_cell_q.value = 0

                proj_norm_value = row_data["proj_norm"]
                ms_norm_value = row_data["ms_norm"]
                lookup_primary_id, lookup_fallback_id = _determine_lookup_ids(proj_norm_value, ms_norm_value)

                # Abrechnungsart cell (column C)
                billing_cell = ws.cell(row=current_row, column=3)
                if lookup_primary_id or lookup_fallback_id:
                    lookup_id = lookup_primary_id if lookup_primary_id else lookup_fallback_id
                    billing_formula = (
                        f'=IFERROR(INDEX(\'Projekt-Budget-Übersicht\'!$C:$C,'
                        f'MATCH({lookup_id},\'Projekt-Budget-Übersicht\'!$L:$L,0)),"")'
                    )
                    billing_cell.value = billing_formula
                else:
                    billing_cell.value = "Unbekannt"

                # Stundensatz cell (column L)
                rate_cell = ws.cell(row=current_row, column=12)
                rate_cell.number_format = '#,##0.00'
                if lookup_primary_id or lookup_fallback_id:
                    lookup_id = lookup_primary_id if lookup_primary_id else lookup_fallback_id
                    rate_formula = (
                        f'=IF($B$2="-",0,'
                        f'IFERROR(INDEX('
                        f'IF($B$2="SV",\'Projekt-Budget-Übersicht\'!$H:$H,'
                        f'IF($B$2="CAD",\'Projekt-Budget-Übersicht\'!$I:$I,'
                        f'IF($B$2="ADM",\'Projekt-Budget-Übersicht\'!$J:$J,'
                        f'\'Projekt-Budget-Übersicht\'!$H:$H))),'
                        f'MATCH({lookup_id},\'Projekt-Budget-Übersicht\'!$L:$L,0)),0))'
                    )
                else:
                    rate_formula = "=IF($B$2=\"-\",0,0)"
                rate_cell.value = rate_formula

                billing_type = (resolved_budget.get("Abrechnungsart", "").strip()
                                if resolved_budget else "")

                # Umsatz cell (column M) - Formula reads Abrechnungsart from column C dynamically
                # Rules: 1) No adjustments (I) in revenue calculation
                #        2) Pauschale: Revenue capped at remaining budget up to 100%
                #        3) Nachweis: ALWAYS full amount (hourly rate × hours worked)
                revenue_cell = ws.cell(row=current_row, column=13)
                revenue_cell.number_format = '#,##0.00'
                # Dynamic formula that checks column C (Abrechnungsart)
                # For Pauschale: MIN(monthly proportional revenue, remaining budget until 100%)
                revenue_formula = (
                    f'=IF($B$2="-",0,'
                    f'IF(C{current_row}="Pauschale",'
                    f'IF(OR(Q{current_row}=0,N(D{current_row})=0),0,'
                    f'MAX(0,MIN(Q{current_row}*((N(F{current_row})+N(K{current_row}))/N(D{current_row})),'
                    f'Q{current_row}-Q{current_row}*(N(E{current_row})/N(D{current_row}))))),'
                    f'L{current_row}*(N(F{current_row})+N(K{current_row}))))'
                )
                revenue_cell.value = revenue_formula

                # Möglicher Umsatz cell (column N) - Same formula but WITHOUT >100% check
                # Shows what COULD be billed (without >100% restriction for Pauschale)
                possible_revenue_cell = ws.cell(row=current_row, column=14)
                possible_revenue_cell.number_format = '#,##0.00'
                possible_revenue_formula = (
                    f'=IF($B$2="-",0,'
                    f'IF(C{current_row}="Pauschale",'
                    f'IF(OR(Q{current_row}=0,N(D{current_row})=0),0,'
                    f'Q{current_row}*((N(F{current_row})+N(K{current_row}))/N(D{current_row}))),'
                    f'L{current_row}*(N(F{current_row})+N(K{current_row}))))'
                )
                possible_revenue_cell.value = possible_revenue_formula

                # Entgangener Umsatz cell (column O) - Difference between Möglicher Umsatz and Umsatz
                lost_revenue_cell = ws.cell(row=current_row, column=15)
                lost_revenue_cell.number_format = '#,##0.00'
                lost_revenue_cell.value = f"=N{current_row}-M{current_row}"

                # Umsatz kumuliert cell (column P) - Placeholder
                budget_earned_cell = ws.cell(row=current_row, column=16)
                budget_earned_cell.number_format = '#,##0.00'
                budget_earned_cell.value = 0

                # Budget Gesamt cell (column Q)
                budget_total_cell = ws.cell(row=current_row, column=17)
                budget_total_cell.number_format = '#,##0.00'
                budget_expr = _build_lookup_expr("E", lookup_primary_id, lookup_fallback_id)
                budget_total_cell.value = f"={budget_expr}"

                # Kosten cell (column R)
                budget_ist_cell = ws.cell(row=current_row, column=18)
                budget_ist_cell.number_format = '#,##0.00'
                if resolved_budget:
                    istkosten_value = resolved_budget.get("Istkosten", 0) or 0
                    budget_ist_cell.value = istkosten_value if istkosten_value != 0 else ""
                else:
                    budget_ist_cell.value = ""

                # Track row for quarterly assignments
                track_key_q = (row_data["proj_norm"], row_data["ms_norm"])
                quarter_row_assignments[track_key_q] = current_row
                quarter_row_assignments_all[emp][track_key_q] = current_row
                revenue_cells_by_key_q.setdefault(track_key_q, []).append((sheet_name, current_row))
                revenue_cells_by_key_q_all.setdefault(track_key_q, []).append((sheet_name, current_row))

                # Track rows for base hour calculation
                if is_special_project:
                    special_project_rows.append(current_row)
                else:
                    regular_project_rows.append(current_row)

                # Color percentage cell
                if should_color:
                    pct_cell = ws.cell(row=current_row, column=7)
                    pct_cell.fill = PatternFill(
                        start_color=status_color_hex(prozent),
                        end_color=status_color_hex(prozent),
                        fill_type="solid",
                    )

                current_row += 1

            # Merge project cells
            block_size = len(proj_block)
            if block_size > 1:
                ws.merge_cells(start_row=block_start, start_column=1,
                               end_row=block_start + block_size - 1, end_column=1)
                ws.cell(row=block_start, column=1).alignment = Alignment(vertical="top")

        # Quarterly summary rows - BASED ON MONTHLY SUMMARY ROWS, NOT PROJECT ROWS
        quarter_data_end_row = current_row - 1

        ws.append([])
        current_row += 1

        # Sum row - Sum of monthly "Summe" rows (total productive hours) + Umsatz sums
        ws.append(["", "Summe", "", "", "", 0, "", "", "", "", "", "", "", "", "", ""])
        sum_row_idx_q = current_row
        for cell in ws[current_row]:
            cell.font = Font(bold=True)
            cell.border = border
        sum_total_cell_q = ws.cell(row=sum_row_idx_q, column=6)
        sum_total_cell_q.number_format = "0.00"
        if monthly_sum_total_cells:
            sum_total_cell_q.value = f"=SUM({','.join(monthly_sum_total_cells)})"
        else:
            sum_total_cell_q.value = 0

        # Add sum formulas for Umsatz (M), Möglicher Umsatz (N), Entgangener Umsatz (O), and Umsatz kumuliert (P) in quarterly summary
        sum_umsatz_cell_q = ws.cell(row=sum_row_idx_q, column=13)  # Column M
        sum_umsatz_cell_q.number_format = '#,##0.00'
        if quarter_data_start_row <= quarter_data_end_row:
            sum_umsatz_cell_q.value = f"=SUM(M{quarter_data_start_row}:M{quarter_data_end_row})"
        else:
            sum_umsatz_cell_q.value = 0

        sum_possible_umsatz_cell_q = ws.cell(row=sum_row_idx_q, column=14)  # Column N
        sum_possible_umsatz_cell_q.number_format = '#,##0.00'
        if quarter_data_start_row <= quarter_data_end_row:
            sum_possible_umsatz_cell_q.value = f"=SUM(N{quarter_data_start_row}:N{quarter_data_end_row})"
        else:
            sum_possible_umsatz_cell_q.value = 0

        sum_lost_umsatz_cell_q = ws.cell(row=sum_row_idx_q, column=15)  # Column O
        sum_lost_umsatz_cell_q.number_format = '#,##0.00'
        if quarter_data_start_row <= quarter_data_end_row:
            sum_lost_umsatz_cell_q.value = f"=SUM(O{quarter_data_start_row}:O{quarter_data_end_row})"
        else:
            sum_lost_umsatz_cell_q.value = 0

        sum_umsatz_kum_cell_q = ws.cell(row=sum_row_idx_q, column=16)  # Column P
        sum_umsatz_kum_cell_q.number_format = '#,##0.00'
        if quarter_data_start_row <= quarter_data_end_row:
            sum_umsatz_kum_cell_q.value = f"=SUM(P{quarter_data_start_row}:P{quarter_data_end_row})"
        else:
            sum_umsatz_kum_cell_q.value = 0

        current_row += 1

        # Bonusberechtigte Stunden row - split into Base (G) + Adjustment (H) = Total (F)
        ws.append(["", "Bonusberechtigte Stunden", "", "", "", 0, 0, 0, "", "", "", "", "", "", "", ""])
        bonus_row_idx_q = current_row
        for cell in ws[current_row]:
            cell.font = Font(bold=True)
            cell.border = border

        # Column G (Basis) - Sum of monthly bonus BASE values (G cells from monthly summaries)
        bonus_base_cell_q = ws.cell(row=bonus_row_idx_q, column=7)
        bonus_base_cell_q.number_format = "0.00"
        if monthly_bonus_base_cells:
            bonus_base_cell_q.value = f"=SUM({','.join(monthly_bonus_base_cells)})"
        else:
            bonus_base_cell_q.value = 0

        # Column H (Anpassung) - Should be 0 because monthly adjustments are already in the monthly F totals
        # The quarterly detail H cells show the monthly adjustments for reference, but we don't sum them again here
        # User can manually enter additional quarterly-level adjustments here if needed
        bonus_adj_cell_q = ws.cell(row=bonus_row_idx_q, column=8)
        bonus_adj_cell_q.number_format = "0.00"
        bonus_adj_cell_q.value = 0  # No automatic summation to avoid double-counting

        # Column F (Total) - Base + Adjustments
        bonus_total_cell_q = ws.cell(row=bonus_row_idx_q, column=6)
        bonus_total_cell_q.number_format = "0.00"
        bonus_total_cell_q.value = f"={bonus_base_cell_q.coordinate}+{bonus_adj_cell_q.coordinate}"
        current_row += 1

        # Bonusberechtigte Stunden Sonderprojekt row - split into Base (G) + Adjustment (H) = Total (F)
        ws.append(["", "Bonusberechtigte Stunden Sonderprojekt", "", "", "", 0, 0, 0, "", "", "", "", "", "", "", ""])
        special_row_idx_q = current_row
        for cell in ws[current_row]:
            cell.font = Font(bold=True)
            cell.border = border

        # Column G (Basis) - Sum of monthly special bonus BASE values (G cells from monthly summaries)
        special_base_cell_q = ws.cell(row=special_row_idx_q, column=7)
        special_base_cell_q.number_format = "0.00"
        if monthly_special_base_cells:
            special_base_cell_q.value = f"=SUM({','.join(monthly_special_base_cells)})"
        else:
            special_base_cell_q.value = 0

        # Column H (Anpassung) - Should be 0 because monthly adjustments are already in the monthly F totals
        # The quarterly detail H cells show the monthly adjustments for reference, but we don't sum them again here
        # User can manually enter additional quarterly-level adjustments here if needed
        special_adj_cell_q = ws.cell(row=special_row_idx_q, column=8)
        special_adj_cell_q.number_format = "0.00"
        special_adj_cell_q.value = 0  # No automatic summation to avoid double-counting

        # Column F (Total) - Base + Adjustments
        special_total_cell_q = ws.cell(row=special_row_idx_q, column=6)
        special_total_cell_q.number_format = "0.00"
        special_total_cell_q.value = f"={special_base_cell_q.coordinate}+{special_adj_cell_q.coordinate}"
        current_row += 1

        # Zugeordnete Stunden von anderen MA (Quartal) - Sum of monthly "Zugeordnete Stunden von anderen MA" rows
        ws.append(["", "Zugeordnete Stunden von anderen MA", "", "", "", 0, "", "", "", "", "", "", "", "", "", ""])
        assigned_row_idx_q = current_row
        for cell in ws[current_row]:
            cell.font = Font(bold=True)
            cell.border = border
        assigned_total_cell_q = ws.cell(row=assigned_row_idx_q, column=6)
        assigned_total_cell_q.number_format = "0.00"
        if monthly_assigned_from_others_cells:
            assigned_total_cell_q.value = f"=SUM({','.join(monthly_assigned_from_others_cells)})"
        else:
            assigned_total_cell_q.value = 0
        current_row += 1

        # Gesamt Bonus Stunden (Quartal) = Bonusberechtigte + Sonderprojekt + Zugeordnete
        ws.append(["", "Gesamt Bonus Stunden", "", "", "", 0, "", "", "", "", "", "", "", "", "", ""])
        total_bonus_row_idx_q = current_row
        for cell in ws[current_row]:
            cell.font = Font(bold=True)
            cell.border = border
            cell.fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
        total_bonus_cell_q = ws.cell(row=total_bonus_row_idx_q, column=6)
        total_bonus_cell_q.number_format = "0.00"
        total_bonus_cell_q.value = f"={bonus_total_cell_q.coordinate}+{special_total_cell_q.coordinate}+{assigned_total_cell_q.coordinate}"
        current_row += 1

        # ========== END QUARTERLY SUMMARY TABLE ==========

        ws.append([])
        current_row += 1
        ws.append([f"--- Gesamtstunden {target_quarter} ---"])
        ws[f"A{current_row}"].font = Font(bold=True, size=12)
        current_row += 1
        ws.append(["Gesamt eingetragene Stunden:", round(total_hours_all_months, 2)])
        for cell in ws[current_row]:
            cell.font = Font(bold=True)
        current_row += 1

        ws.append(["Bonusberechtigte Stunden (Quartal):", 0])
        quarter_bonus_row = current_row
        quarter_bonus_cell = ws.cell(row=quarter_bonus_row, column=2)
        if monthly_bonus_total_cells:
            quarter_bonus_cell.value = f"=SUM({','.join(monthly_bonus_total_cells)})"
        else:
            quarter_bonus_cell.value = round(total_bonus_hours_quarter, 2)
        quarter_bonus_cell.number_format = "0.00"
        for cell in ws[current_row]:
            cell.font = Font(bold=True)
        current_row += 1

        ws.append(["Bonusberechtigte Stunden Sonderprojekt (Quartal):", 0])
        quarter_special_row = current_row
        quarter_special_cell = ws.cell(row=quarter_special_row, column=2)
        if monthly_special_bonus_total_cells:
            quarter_special_cell.value = f"=SUM({','.join(monthly_special_bonus_total_cells)})"
        else:
            quarter_special_cell.value = round(total_bonus_special_hours_quarter, 2)
        quarter_special_cell.number_format = "0.00"
        for cell in ws[current_row]:
            cell.font = Font(bold=True)
        current_row += 1

        # Store quarterly summary cell references
        employee_summary_data[emp]['quarter_total_hours_cell'] = f"'{sheet_name}'!B{quarter_bonus_row - 1}"
        employee_summary_data[emp]['quarter_bonus_hours_cell'] = f"'{sheet_name}'!{quarter_bonus_cell.coordinate}"
        employee_summary_data[emp]['quarter_special_bonus_hours_cell'] = f"'{sheet_name}'!{quarter_special_cell.coordinate}"

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 18  # Abrechnungsart
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 12  # Differenz
        ws.column_dimensions['J'].width = 25  # Zuordnen an (Dropdown)
        ws.column_dimensions['K'].width = 15  # Von anderen
        ws.column_dimensions['L'].width = 18  # Stundensatz (€/h)
        ws.column_dimensions['M'].width = 15  # Umsatz (€)
        ws.column_dimensions['N'].width = 18  # Möglicher Umsatz (€)
        ws.column_dimensions['O'].width = 18  # Entgangener Umsatz (€)
        ws.column_dimensions['P'].width = 22  # Umsatz kumuliert (€)
        ws.column_dimensions['Q'].width = 22  # Soll Obermeilenstein (h)
        ws.column_dimensions['R'].width = 18  # Budget Gesamt (€)
        ws.column_dimensions['S'].width = 15  # Kosten (€)
        ws.column_dimensions['T'].width = 12  # Rechnung
        ws.column_dimensions['U'].width = 30  # Kommentar

        progress = int((idx_emp / total_emps) * 80) + 20
        progress_cb(min(progress, 95), f"Verarbeite Mitarbeiter {emp}")

    # Second pass: Fill "Von anderen" formulas
    progress_cb(90, "Erstelle Zuordnungs-Formeln")
    for emp in employees:
        ws = wb[emp[:31]]
        for track_key, row_num in row_assignments[emp].items():
            proj_norm, ms_norm, month = track_key
            # Build formula to sum hours from other employees who assigned to this employee
            formula_parts = []
            for other_emp in employees:
                if other_emp == emp:
                    continue
                other_key = (proj_norm, ms_norm, month)
                if other_key in row_assignments[other_emp]:
                    other_row = row_assignments[other_emp][other_key]
                    other_sheet = other_emp[:31]
                    # Add SUMIF formula part: IF assign cell (J) = current emp, then add diff cell (I)
                    formula_parts.append(f"IF('{other_sheet}'!J{other_row}=\"{emp}\",'{other_sheet}'!I{other_row},0)")

            # Set the formula in "Von anderen" cell (column K)
            from_others_cell = ws.cell(row=row_num, column=11)
            if formula_parts:
                formula = "=" + "+".join(formula_parts)
                from_others_cell.value = formula
            else:
                from_others_cell.value = 0

            # Set cumulative revenue formula in column P
            # Sum ALL revenue for this project/milestone in the current month, regardless of position
            revenue_total_cell = ws.cell(row=row_num, column=16)
            revenue_total_cell.number_format = '#,##0.00'
            revenue_refs = revenue_cells_by_key.get(track_key, [])
            if revenue_refs:
                parts = [
                    f"'{sheet}'!M{rev_row}"
                    for sheet, rev_row in revenue_refs
                ]
                revenue_total_cell.value = "=" + "+".join(parts)
            else:
                revenue_total_cell.value = 0

        # Fill "Zugeordnete Stunden von anderen MA" sum formulas
        for month, (start_row, end_row, assigned_row) in month_sections[emp].items():
            assigned_cell = ws.cell(row=assigned_row, column=6)
            # Sum all "Von anderen (K)" cells in this month section
            if start_row <= end_row:
                assigned_cell.value = f"=SUM(K{start_row}:K{end_row})"
            else:
                assigned_cell.value = 0

        # Fill quarterly "Quartal (h)", "Von anderen" and "Umsatz kumuliert" formulas
        for track_key_q, row_num in quarter_row_assignments_all[emp].items():
            proj_norm, ms_norm = track_key_q

            # Quartal (h) (column F) - Sum of (F + I + K) from monthly tables
            # This ensures the quarterly total reflects manual adjustments and assignments
            quarter_hours_cell = ws.cell(row=row_num, column=6)
            quarter_hours_cell.number_format = "0.00"
            monthly_f_refs = []
            monthly_i_refs = []
            monthly_k_refs = []
            for month in months:
                monthly_key = (proj_norm, ms_norm, month)
                if monthly_key in row_assignments[emp]:
                    monthly_row = row_assignments[emp][monthly_key]
                    monthly_f_refs.append(f"F{monthly_row}")
                    monthly_i_refs.append(f"I{monthly_row}")
                    monthly_k_refs.append(f"K{monthly_row}")

            if monthly_f_refs:
                # Sum all monthly F+I+K values
                all_refs = monthly_f_refs + monthly_i_refs + monthly_k_refs
                quarter_hours_cell.value = f"=SUM({','.join(all_refs)})"
            else:
                quarter_hours_cell.value = 0

            # Von anderen (column K) - Sum hours from ALL monthly tables for this employee
            # across all months where this project/milestone appears
            from_others_cell_q = ws.cell(row=row_num, column=11)
            monthly_from_others_refs = []
            for month in months:
                monthly_key = (proj_norm, ms_norm, month)
                if monthly_key in row_assignments[emp]:
                    monthly_row = row_assignments[emp][monthly_key]
                    monthly_from_others_refs.append(f"K{monthly_row}")

            if monthly_from_others_refs:
                from_others_cell_q.value = f"=SUM({','.join(monthly_from_others_refs)})"
            else:
                from_others_cell_q.value = 0

            # Umsatz kumuliert (column P) - Sum ALL revenue for this project/milestone across ALL employees
            revenue_total_cell_q = ws.cell(row=row_num, column=16)
            revenue_total_cell_q.number_format = '#,##0.00'
            revenue_refs_q = revenue_cells_by_key_q_all.get(track_key_q, [])
            if revenue_refs_q:
                parts = [
                    f"'{sheet}'!M{rev_row}"
                    for sheet, rev_row in revenue_refs_q
                ]
                revenue_total_cell_q.value = "=" + "+".join(parts)
            else:
                revenue_total_cell_q.value = 0

    # Create summary cover sheet
    progress_cb(96, "Erstelle Deckblatt")
    _create_cover_sheet(
        wb, target_quarter, months, employee_summary_data, border, report_title=report_title
    )

    # Always save macro workbooks with .xlsm when template/VBA is involved
    save_path = out_path
    if (used_template or add_vba) and out_path.suffix.lower() != ".xlsm":
        save_path = out_path.with_suffix(".xlsm")

    save_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(save_path)

    # Add VBA macro if requested (kept for backwards compatibility)
    final_path = save_path
    if add_vba and not used_template:
        progress_cb(98, "Füge VBA-Makro hinzu")
        final_path = _add_vba_macro(save_path, progress_cb)
    else:
        progress_cb(98, "Speichere .xlsm (Makro enthalten)" if used_template else "Überspringe VBA-Makro")

    return final_path


def generate_quarterly_report(
    csv_path: Path,
    xml_path: Path,
    output_dir: Path,
    output_name_prefix: Optional[str] = None,
    requested_quarter: Optional[str] = None,
    progress_cb: ProgressCallback = _noop_progress,
) -> Path:
    """Hauptfunktion: erzeugt den Bericht und gibt den Pfad zur Excel-Datei zurück."""

    progress_cb(5, "Lade CSV-Daten")
    df_csv = load_csv_projects(csv_path)

    progress_cb(8, "Lade Budget-Daten")
    df_budget, milestone_parent_map = load_csv_budget_data(csv_path)

    progress_cb(10, "Lade XML-Daten")
    df_xml = load_xml_times(xml_path)

    selection = determine_quarter(df_xml, requested=requested_quarter)
    progress_cb(15, f"Wähle Quartal {selection.period}")

    year = selection.period.year
    quarter_num = selection.period.quarter
    # Output is now .xlsm because we use the template
    base_name = f"Q{quarter_num}-{year}.xlsm"
    prefix = (output_name_prefix or "").strip()
    if prefix:
        base_name = f"{prefix}_{base_name}"
    out_path = output_dir / base_name

    result = build_quarterly_report(
        df_csv=df_csv,
        df_budget=df_budget,
        milestone_parent_map=milestone_parent_map,
        df_xml=df_xml,
        target_quarter=selection.period,
        months=selection.months,
        out_path=out_path,
        progress_cb=progress_cb,
    )

    progress_cb(100, "Fertig")
    return result


__all__ = [
    "generate_quarterly_report",
    "determine_quarter",
    "list_available_quarters",
    "parse_quarter",
    "MONTHLY_BUDGETS",
    "QUARTERLY_BUDGETS",
]
